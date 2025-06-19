# log_expense.py

"""
log_expense.py - Tool for logging expenses against budget allocations in CityInfraXLS

This script prompts for expense details, validates against existing budget allocations,
assigns a unique expense ID, records the expense, and updates the remaining budget
in the allocated budget entry. Both operations are handled atomically.
"""

import os
import json
import argparse
import datetime
import uuid
import tempfile
import shutil
from decimal import Decimal, InvalidOperation
import pandas as pd
from openpyxl import load_workbook
from utils.excel_handler import create_sheets_from_schema, load_workbook, save_workbook

def generate_expense_id():
    """Generate a unique expense ID with EXP prefix."""
    # Take first 8 characters of a UUID and prefix with EXP
    unique_id = str(uuid.uuid4()).replace('-', '')[:8].upper()
    return f"EXP-{unique_id}"

def load_departments():
    """Load list of departments from budget allocations sheet."""
    budget_path = "data/budget_allocations.xlsx"
    if not os.path.exists(budget_path):
        raise FileNotFoundError(f"Budget allocation file not found: {budget_path}")
    
    try:
        # Load budget allocations
        wb = load_workbook(budget_path)
        ws = wb["Allocations"]
        
        # Get column index for department
        headers = [cell.value for cell in ws[1]]
        dept_index = headers.index("department") + 1  # +1 because openpyxl is 1-indexed
        
        # Extract unique departments
        departments = set()
        for row in ws.iter_rows(min_row=2):  # Skip header row
            dept = row[dept_index-1].value  # -1 to convert back to 0-indexed
            if dept:
                departments.add(dept)
        
        return list(departments)
    except Exception as e:
        raise Exception(f"Error loading departments: {str(e)}")

def get_budget_info(department, category):
    """
    Get budget allocation information for a department and category.
    
    Args:
        department (str): Department name
        category (str): Expense category
    
    Returns:
        tuple: (budget_row, project_id, fiscal_year, allocated_amount, spent_amount, remaining_budget)
    """
    budget_path = "data/budget_allocations.xlsx"
    expense_path = "data/expenses.xlsx"
    
    # Load budget allocations
    budget_df = pd.read_excel(budget_path, sheet_name="Allocations")
    
    # Filter by department and category
    dept_budgets = budget_df[(budget_df["department"] == department) & 
                            (budget_df["category"] == category) &
                            (budget_df["status"].isin(["approved", "allocated"]))]
    
    if dept_budgets.empty:
        return None, None, None, 0, 0, 0
    
    # Get latest fiscal year budget for this department/category
    latest_budget = dept_budgets.sort_values("allocation_date", ascending=False).iloc[0]
    budget_row = latest_budget.name  # Get the index of the row
    project_id = latest_budget["project_id"]
    fiscal_year = latest_budget["fiscal_year"]
    allocated = float(latest_budget["allocated_amount"])
    
    # Get spent amount from existing expenses if available
    if os.path.exists(expense_path):
        expense_df = pd.read_excel(expense_path, sheet_name="Expenses")
        
        # Filter expenses by project_id
        project_expenses = expense_df[expense_df["project_id"] == project_id]
        
        # Sum up expenses
        total_spent = project_expenses["amount"].sum() if not project_expenses.empty else 0
    else:
        total_spent = 0
    
    # Calculate remaining budget
    remaining = allocated - total_spent
    
    return budget_row, project_id, fiscal_year, allocated, total_spent, remaining

def prompt_for_category():
    """Prompt user for expense category."""
    categories = ["maintenance", "new_construction", "renovation", 
                  "emergency", "planning", "other"]
    
    print("Available categories:")
    for i, category in enumerate(categories, 1):
        print(f"{i}. {category}")
    
    while True:
        choice = input("Select category (number): ").strip()
        try:
            index = int(choice) - 1
            if 0 <= index < len(categories):
                return categories[index]
            else:
                print("Invalid selection. Please try again.")
        except ValueError:
            print("Please enter a number.")

def create_expense_schema():
    """Create a temporary schema file for expenses."""
    expense_schema = {
        "type": "object",
        "properties": {
            "expense_id": {"type": "string"},
            "project_id": {"type": "string"},
            "department": {"type": "string"},
            "amount": {"type": "number"},
            "category": {"type": "string"},
            "description": {"type": "string"},
            "date": {"type": "string", "format": "date"},
            "fiscal_year": {"type": "string"},
            "recorded_by": {"type": "string"},
            "recorded_on": {"type": "string", "format": "date-time"},
            "remaining_budget": {"type": "number"}
        },
        "required": ["expense_id", "project_id", "department", "amount", 
                     "category", "description", "date", "fiscal_year"]
    }
    
    # Write schema to temporary file
    temp_schema_path = "temp_expense_schema.json"
    with open(temp_schema_path, 'w') as f:
        json.dump(expense_schema, f, indent=2)
    
    return temp_schema_path

def update_budget_and_log_expense(expense_data, budget_row, new_spent_amount, new_remaining_budget):
    """
    Update budget allocation and log expense atomically.
    
    Args:
        expense_data (dict): Expense data to log
        budget_row (int): Row index of the budget to update
        new_spent_amount (float): New spent amount after the expense
        new_remaining_budget (float): New remaining budget after the expense
    """
    budget_path = "data/budget_allocations.xlsx"
    expense_path = "data/expenses.xlsx"
    
    # Create temporary files
    temp_budget_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_budget_path = temp_budget_file.name
    temp_budget_file.close()
    
    temp_expense_file = None
    temp_expense_path = None
    if os.path.exists(expense_path):
        temp_expense_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_expense_path = temp_expense_file.name
        temp_expense_file.close()
    
    try:
        # 1. Update budget allocation
        shutil.copy2(budget_path, temp_budget_path)
        budget_wb = load_workbook(temp_budget_path)
        budget_ws = budget_wb["Allocations"]
        
        # Get header row to find column indices
        headers = [cell.value for cell in budget_ws[1]]
        
        # Check if spent_amount and remaining_budget columns exist
        if "spent_amount" not in headers:
            # Add spent_amount and remaining_budget columns if they don't exist
            budget_ws.cell(1, len(headers) + 1).value = "spent_amount"
            budget_ws.cell(1, len(headers) + 2).value = "remaining_budget"
            headers.extend(["spent_amount", "remaining_budget"])
        
        # Get column indices
        spent_col_idx = headers.index("spent_amount") + 1
        remaining_col_idx = headers.index("remaining_budget") + 1
        
        # Update the row (budget_row is 0-indexed in pandas, but 1-indexed in openpyxl and we need to account for header)
        actual_row = budget_row + 2  # +2 because pandas index is 0-based and we need to skip header row
        
        budget_ws.cell(actual_row, spent_col_idx).value = new_spent_amount
        budget_ws.cell(actual_row, remaining_col_idx).value = new_remaining_budget
        
        # Flag negative balance with cell formatting
        if new_remaining_budget < 0:
            from openpyxl.styles import Font, PatternFill
            red_font = Font(color="FF0000", bold=True)
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            budget_ws.cell(actual_row, remaining_col_idx).font = red_font
            budget_ws.cell(actual_row, remaining_col_idx).fill = red_fill
        
        # Save budget workbook
        save_workbook(budget_wb, temp_budget_path)
        
        # 2. Append expense to expense tracker
        # Add remaining budget to expense data
        expense_data["remaining_budget"] = new_remaining_budget
        
        if os.path.exists(expense_path):
            # Copy existing file
            shutil.copy2(expense_path, temp_expense_path)
            expense_wb = load_workbook(temp_expense_path)
        else:
            # Create new expense file
            temp_schema_path = create_expense_schema()
            create_sheets_from_schema(temp_schema_path, temp_expense_path, "Expenses")
            os.remove(temp_schema_path)
            expense_wb = load_workbook(temp_expense_path)
        
        expense_ws = expense_wb["Expenses"]
        
        # Get headers
        headers = [cell.value for cell in expense_ws[1]]
        
        # Create new row
        new_row = []
        for header in headers:
            new_row.append(expense_data.get(header, ""))
        
        # Append row
        expense_ws.append(new_row)
        
        # Save expense workbook
        save_workbook(expense_wb, temp_expense_path)
        
        # 3. If both operations succeeded, replace the original files
        shutil.move(temp_budget_path, budget_path)
        if temp_expense_path:
            shutil.move(temp_expense_path, expense_path)
        else:
            # Just in case temp_expense_path wasn't created but we need the expense file
            if not os.path.exists(expense_path):
                os.makedirs(os.path.dirname(expense_path), exist_ok=True)
                shutil.copy2(temp_expense_path, expense_path)
        
        print(f"Expense recorded and budget updated successfully.")
        
    except Exception as e:
        # Clean up temporary files
        if os.path.exists(temp_budget_path):
            os.unlink(temp_budget_path)
        if temp_expense_path and os.path.exists(temp_expense_path):
            os.unlink(temp_expense_path)
        
        raise Exception(f"Failed to update budget and log expense: {str(e)}")

def main():
    parser = argparse.ArgumentParser(description="Log an expense against a budget allocation")
    parser.add_argument("--batch", action="store_true", help="Run in batch mode without prompts")
    parser.add_argument("--force", action="store_true", help="Force expense logging even if budget is exceeded")
    args = parser.parse_args()
    
    try:
        # Load departments
        departments = load_departments()
        
        if not departments:
            print("No budget allocations found. Please allocate a budget first.")
            return 1
        
        # Initialize expense data
        expense_data = {
            "expense_id": generate_expense_id(),
            "recorded_on": datetime.datetime.now().isoformat()
        }
        
        # Department
        if not args.batch:
            print("Available departments:")
            for i, dept in enumerate(departments, 1):
                print(f"{i}. {dept}")
            
            while True:
                choice = input("Select department (number): ").strip()
                try:
                    index = int(choice) - 1
                    if 0 <= index < len(departments):
                        expense_data["department"] = departments[index]
                        break
                    else:
                        print("Invalid selection. Please try again.")
                except ValueError:
                    print("Please enter a number.")
        else:
            expense_data["department"] = input().strip()
            if expense_data["department"] not in departments:
                raise ValueError(f"Department not found: {expense_data['department']}")
        
        # Expense category
        expense_data["category"] = prompt_for_category()
        
        # Get budget information
        budget_row, project_id, fiscal_year, allocated, spent, remaining = get_budget_info(
            expense_data["department"], 
            expense_data["category"]
        )
        
        if budget_row is None:
            print(f"No budget allocation found for {expense_data['department']} in category {expense_data['category']}.")
            if not args.force:
                proceed = input("Do you want to proceed anyway? (y/n): ").strip().lower()
                if proceed != 'y':
                    print("Operation canceled.")
                    return 1
        else:
            print(f"Current budget status for {expense_data['department']} in {expense_data['category']}:")
            print(f"  Allocated: ${allocated:.2f}")
            print(f"  Spent: ${spent:.2f}")
            print(f"  Remaining: ${remaining:.2f}")
            expense_data["project_id"] = project_id
            expense_data["fiscal_year"] = fiscal_year
        
        # Expense date
        while True:
            date_str = input("Expense date (YYYY-MM-DD, leave empty for today): ").strip()
            if not date_str:
                expense_data["date"] = datetime.date.today().isoformat()
                break
            
            try:
                # Validate date format
                expense_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
                expense_data["date"] = expense_date.isoformat()
                break
            except ValueError:
                print("Invalid date format. Please use YYYY-MM-DD.")
        
        # Expense amount
        while True:
            try:
                amount_str = input("Expense amount ($): ").strip()
                amount = float(Decimal(amount_str))
                if amount <= 0:
                    print("Amount must be positive.")
                    continue
                expense_data["amount"] = amount
                break
            except (ValueError, InvalidOperation):
                print("Please enter a valid number.")
        
        # Check budget
        if budget_row is not None:
            new_spent = spent + expense_data["amount"]
            new_remaining = remaining - expense_data["amount"]
            
            if new_remaining < 0 and not args.force:
                print(f"Warning: This expense will exceed the budget by ${abs(new_remaining):.2f}")
                proceed = input("Do you want to proceed anyway? (y/n): ").strip().lower()
                if proceed != 'y':
                    print("Operation canceled.")
                    return 1
        else:
            # No budget found
            new_spent = expense_data["amount"]
            new_remaining = -expense_data["amount"]
        
        # Expense description
        while True:
            description = input("Expense description: ").strip()
            if len(description) < 5:
                print("Description must be at least 5 characters.")
                continue
            expense_data["description"] = description
            break
        
        # Ask for recorded_by
        recorded_by = input("Recorded by (optional): ").strip()
        if recorded_by:
            expense_data["recorded_by"] = recorded_by
        
        # Update budget and log expense atomically
        update_budget_and_log_expense(expense_data, budget_row, new_spent, new_remaining)
        
        print(f"Expense {expense_data['expense_id']} for ${expense_data['amount']:.2f} "
              f"has been recorded for {expense_data['department']}.")
        
        if new_remaining < 0:
            print(f"WARNING: Budget exceeded by ${abs(new_remaining):.2f}")
        
    except KeyboardInterrupt:
        print("\nOperation canceled.")
        return 1
    except Exception as e:
        print(f"Error: {str(e)}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())