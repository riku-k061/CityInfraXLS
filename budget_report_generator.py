# budget_report_generator.py

"""
budget_report_generator.py - Generate comprehensive budget reports in CityInfraXLS

This module creates a consolidated budget report workbook with multiple sheets:
- Summary: Overview of all departments with allocation vs. expense totals
- Department Details: Detailed transaction history by department
- Alerts: Budget warnings, overruns, and other financial flags
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.worksheet.table import Table, TableStyleInfo

def format_currency(amount):
    """Format amount as currency string."""
    if pd.isna(amount):
        return "$0.00"
    return f"${float(amount):.2f}"

def generate_budget_report(output_path=None, fiscal_year=None):
    """
    Generate a comprehensive budget report with multiple sheets.
    
    Args:
        output_path (str): Path where the report will be saved
        fiscal_year (str): Fiscal year to filter data (YYYY-YYYY format)
                          If None, latest fiscal year will be used
    
    Returns:
        str: Path to the generated report
    """
    # Set default output path if not provided
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = "reports"
        os.makedirs(output_dir, exist_ok=True)
        output_path = f"{output_dir}/budget_report_{timestamp}.xlsx"
    
    # Load data
    budget_df = pd.read_excel("data/budget_allocations.xlsx", sheet_name="Allocations")
    expenses_df = pd.read_excel("data/expenses.xlsx", sheet_name="Expenses")
    
    # Determine fiscal year if not provided
    if fiscal_year is None and not budget_df.empty:
        fiscal_year = budget_df["fiscal_year"].mode().iloc[0]
    
    print(f"Generating budget report for fiscal year: {fiscal_year}")
    
    # Filter data by fiscal year if provided
    if fiscal_year:
        budget_df = budget_df[budget_df["fiscal_year"] == fiscal_year]
        expenses_df = expenses_df[expenses_df["fiscal_year"] == fiscal_year]
    
    # Check if we have data
    if budget_df.empty:
        raise ValueError("No budget allocations found for the specified fiscal year")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    summary_sheet = wb.create_sheet("Summary")
    details_sheet = wb.create_sheet("Department Details")
    alerts_sheet = wb.create_sheet("Alerts")
    
    # ============ POPULATE SUMMARY SHEET ============
    
    # Add title
    summary_sheet.merge_cells('A1:H1')
    title_cell = summary_sheet['A1']
    title_cell.value = f"Budget Summary Report - Fiscal Year {fiscal_year}"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Create summary data by department
    summary_data = []
    
    # Group budget data by department
    dept_summary = budget_df.groupby("department").agg({
        "allocated_amount": "sum",
    }).reset_index()
    
    # Add spent amount from expenses
    if not expenses_df.empty:
        dept_expenses = expenses_df.groupby("department").agg({
            "amount": "sum"
        }).reset_index().rename(columns={"amount": "spent_amount"})
        
        # Merge with dept_summary
        dept_summary = pd.merge(dept_summary, dept_expenses, 
                               on="department", how="left").fillna(0)
    else:
        dept_summary["spent_amount"] = 0
    
    # Calculate remaining and percentage
    dept_summary["remaining_amount"] = dept_summary["allocated_amount"] - dept_summary["spent_amount"]
    dept_summary["percent_used"] = (dept_summary["spent_amount"] / dept_summary["allocated_amount"]) * 100
    dept_summary["status"] = np.where(dept_summary["remaining_amount"] < 0, "OVERRUN",
                            np.where(dept_summary["remaining_amount"] == 0, "DEPLETED",
                            np.where(dept_summary["percent_used"] > 90, "LOW", "ACTIVE")))
    
    # Add totals row
    totals = pd.DataFrame({
        "department": ["TOTAL"],
        "allocated_amount": [dept_summary["allocated_amount"].sum()],
        "spent_amount": [dept_summary["spent_amount"].sum()],
        "remaining_amount": [dept_summary["allocated_amount"].sum() - dept_summary["spent_amount"].sum()],
        "percent_used": [(dept_summary["spent_amount"].sum() / dept_summary["allocated_amount"].sum()) * 100 
                          if dept_summary["allocated_amount"].sum() > 0 else 0],
        "status": [""]
    })
    
    dept_summary = pd.concat([dept_summary, totals])
    
    # Write summary table to sheet
    summary_sheet.append(["Department", "Allocated", "Spent", "Remaining", "% Used", "Status"])
    
    for index, row in dept_summary.iterrows():
        formatted_row = [
            row["department"],
            format_currency(row["allocated_amount"]),
            format_currency(row["spent_amount"]),
            format_currency(row["remaining_amount"]),
            f"{row['percent_used']:.1f}%" if not pd.isna(row['percent_used']) else "0.0%",
            row["status"]
        ]
        summary_sheet.append(formatted_row)
    
    # Style header row
    for cell in summary_sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
    # Style the data - conditional formatting
    for row_idx, row in enumerate(summary_sheet.iter_rows(min_row=3, max_row=3+len(dept_summary)-1), 3):
        # Format department totals row
        if row[0].value == "TOTAL":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        else:
            # Format status cell
            status_cell = row[5]
            if status_cell.value == "OVERRUN":
                status_cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                status_cell.font = Font(color="9C0006")
            elif status_cell.value == "LOW":
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(color="9C5700")
            
            # Format remaining amount
            remaining_cell = row[3]
            amount_str = remaining_cell.value.replace("$", "").replace(",", "")
            try:
                if float(amount_str) < 0:
                    remaining_cell.font = Font(color="9C0006")
            except ValueError:
                pass
    
    # Add chart - Budget Allocation by Department
    chart1 = PieChart()
    chart1.title = "Budget Allocation by Department"
    
    # Data for chart
    data = Reference(summary_sheet, min_col=2, min_row=2, max_row=len(dept_summary)+1)
    cats = Reference(summary_sheet, min_col=1, min_row=3, max_row=len(dept_summary))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    
    # Add chart to sheet
    summary_sheet.add_chart(chart1, "I3")
    
    # Add bar chart - Spent vs Remaining
    chart2 = BarChart()
    chart2.title = "Spent vs Remaining by Department"
    chart2.style = 10
    chart2.type = "col"
    chart2.grouping = "stacked"
    
    data = Reference(summary_sheet, min_col=3, min_row=2, max_col=4, max_row=len(dept_summary)+1)
    cats = Reference(summary_sheet, min_col=1, min_row=3, max_row=len(dept_summary))
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    
    # Add chart to sheet
    summary_sheet.add_chart(chart2, "I18")
    
    # Set column widths
    for col_idx, column in enumerate(summary_sheet.columns, 1):
        column_width = 15  # default width
        if col_idx == 1:  # Department column
            column_width = 20
        summary_sheet.column_dimensions[get_column_letter(col_idx)].width = column_width
    
    # ============ POPULATE DEPARTMENT DETAILS SHEET ============
    
    # Add title
    details_sheet.merge_cells('A1:G1')
    title_cell = details_sheet['A1']
    title_cell.value = "Department Budget Details"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # For each department, add a section with budget and expenses
    row_idx = 3
    
    for dept in budget_df["department"].unique():
        # Department header
        details_sheet.merge_cells(f'A{row_idx}:G{row_idx}')
        dept_cell = details_sheet[f'A{row_idx}']
        dept_cell.value = f"Department: {dept}"
        dept_cell.font = Font(size=14, bold=True)
        dept_cell.alignment = Alignment(horizontal='left')
        
        row_idx += 2
        
        # Budget allocations for this department
        dept_budget = budget_df[budget_df["department"] == dept]
        
        details_sheet.merge_cells(f'A{row_idx}:G{row_idx}')
        budget_header = details_sheet[f'A{row_idx}']
        budget_header.value = "Budget Allocations"
        budget_header.font = Font(bold=True)
        budget_header.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        row_idx += 1
        
        # Budget headers
        budget_columns = ["Project ID", "Category", "Allocation Date", "Allocated Amount", 
                         "Spent Amount", "Remaining", "Status"]
        details_sheet.append(budget_columns)
        
        # Style header row
        for cell in details_sheet[row_idx]:
            cell.font = Font(bold=True)
        
        row_idx += 1
        
        # Budget data
        for _, budget_row in dept_budget.iterrows():
            project_id = budget_row["project_id"]
            
            # Calculate spent amount for this project
            spent = 0
            if not expenses_df.empty:
                project_expenses = expenses_df[expenses_df["project_id"] == project_id]
                spent = project_expenses["amount"].sum() if not project_expenses.empty else 0
                
            remaining = float(budget_row["allocated_amount"]) - spent
            
            # Determine status
            if remaining < 0:
                status = "OVERRUN"
            elif remaining == 0:
                status = "DEPLETED"
            elif remaining < (0.1 * float(budget_row["allocated_amount"])):
                status = "LOW"
            else:
                status = "ACTIVE"
                
            details_sheet.append([
                project_id,
                budget_row["category"],
                budget_row["allocation_date"],
                format_currency(budget_row["allocated_amount"]),
                format_currency(spent),
                format_currency(remaining),
                status
            ])
            
            # Style status cell
            status_cell = details_sheet.cell(row=row_idx, column=7)
            if status == "OVERRUN":
                status_cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                status_cell.font = Font(color="9C0006")
            elif status == "LOW":
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(color="9C5700")
            
            row_idx += 1
        
        row_idx += 2
        
        # Expenses for this department
        dept_expenses = expenses_df[expenses_df["department"] == dept]
        
        if not dept_expenses.empty:
            details_sheet.merge_cells(f'A{row_idx}:G{row_idx}')
            exp_header = details_sheet[f'A{row_idx}']
            exp_header.value = "Expense Transactions"
            exp_header.font = Font(bold=True)
            exp_header.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            row_idx += 1
            
            # Expense headers
            expense_columns = ["Expense ID", "Project ID", "Date", "Category", 
                              "Amount", "Description", "Recorded By"]
            details_sheet.append(expense_columns)
            
            # Style header row
            for cell in details_sheet[row_idx]:
                cell.font = Font(bold=True)
            
            row_idx += 1
            
            # Expense data
            for _, expense_row in dept_expenses.iterrows():
                details_sheet.append([
                    expense_row["expense_id"],
                    expense_row["project_id"],
                    expense_row["date"],
                    expense_row["category"],
                    format_currency(expense_row["amount"]),
                    expense_row["description"],
                    expense_row["recorded_by"] if "recorded_by" in expense_row else ""
                ])
                row_idx += 1
        
        row_idx += 2
    
    # Set column widths
    for col_idx, column in enumerate(details_sheet.columns, 1):
        column_width = 15  # default width
        if col_idx == 6:  # Description column
            column_width = 40
        details_sheet.column_dimensions[get_column_letter(col_idx)].width = column_width
    
    # ============ POPULATE ALERTS SHEET ============
    
    # Add title
    alerts_sheet.merge_cells('A1:E1')
    title_cell = alerts_sheet['A1']
    title_cell.value = "Budget Alerts and Warnings"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    row_idx = 3
    alerts_sheet.append(["Department", "Project ID", "Category", "Issue", "Details"])
    
    # Style header row
    for cell in alerts_sheet[row_idx]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    row_idx += 1
    alerts_found = False
    
    # Check for overruns and low budget
    for dept in budget_df["department"].unique():
        dept_budget = budget_df[budget_df["department"] == dept]
        
        for _, budget_row in dept_budget.iterrows():
            project_id = budget_row["project_id"]
            allocated = float(budget_row["allocated_amount"])
            
            # Calculate spent amount
            spent = 0
            if not expenses_df.empty:
                project_expenses = expenses_df[expenses_df["project_id"] == project_id]
                spent = project_expenses["amount"].sum() if not project_expenses.empty else 0
                
            remaining = allocated - spent
            
            # Check for issues
            if remaining < 0:
                alerts_sheet.append([
                    dept,
                    project_id,
                    budget_row["category"],
                    "BUDGET OVERRUN",
                    f"Allocated: {format_currency(allocated)}, Spent: {format_currency(spent)}, " +
                    f"Overrun: {format_currency(abs(remaining))}"
                ])
                # Style this row
                for cell in alerts_sheet[row_idx]:
                    cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                alerts_sheet.cell(row=row_idx, column=4).font = Font(bold=True, color="9C0006")
                row_idx += 1
                alerts_found = True
                
            elif remaining < (0.1 * allocated):
                alerts_sheet.append([
                    dept,
                    project_id,
                    budget_row["category"],
                    "LOW BUDGET",
                    f"Allocated: {format_currency(allocated)}, Spent: {format_currency(spent)}, " +
                    f"Remaining: {format_currency(remaining)} ({remaining/allocated*100:.1f}%)"
                ])
                # Style this row
                for cell in alerts_sheet[row_idx]:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                alerts_sheet.cell(row=row_idx, column=4).font = Font(bold=True, color="9C5700")
                row_idx += 1
                alerts_found = True
    
    # Check for inactive budgets (no expenses for 90+ days)
    if not alerts_found:
        alerts_sheet.append(["No budget alerts found at this time."])
    
    # Set column widths
    for col_idx, column in enumerate(alerts_sheet.columns, 1):
        column_width = 15  # default width
        if col_idx == 5:  # Details column
            column_width = 40
        alerts_sheet.column_dimensions[get_column_letter(col_idx)].width = column_width
    
    # Add hyperlinks between sheets for easy navigation
    # Link from Summary to Department Details
    summary_sheet['A2'].hyperlink = f"#{details_sheet.title}!A1"
    summary_sheet['A2'].style = "Hyperlink"
    
    # Link from Summary to Alerts
    if alerts_found:
        summary_sheet.cell(row=2, column=6).hyperlink = f"#{alerts_sheet.title}!A1"
        summary_sheet.cell(row=2, column=6).style = "Hyperlink"
    
    # Save workbook
    wb.save(output_path)
    print(f"Budget report saved to: {output_path}")
    
    return output_path


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate budget report")
    parser.add_argument("--fiscal_year", type=str, help="Fiscal year (YYYY-YYYY)")
    parser.add_argument("--output", type=str, help="Output file path")
    
    args = parser.parse_args()
    
    try:
        report_path = generate_budget_report(
            output_path=args.output,
            fiscal_year=args.fiscal_year
        )
        print(f"Report generated successfully: {report_path}")
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        exit(1)