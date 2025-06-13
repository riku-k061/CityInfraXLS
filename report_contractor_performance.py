# report_contractor_performance.py

import os
import sys
import pandas as pd
import numpy as np
import argparse
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, Reference

# Add parent directory to path to import modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.excel_handler import create_tasks_sheet

def load_data():
    """Load data from tasks.xlsx, contractors.xlsx, and incidents.xlsx."""
    data_dir = "data"
    files = {
        "tasks": os.path.join(data_dir, "tasks.xlsx"),
        "contractors": os.path.join(data_dir, "contractors.xlsx"),
        "incidents": os.path.join(data_dir, "incidents.xlsx")
    }
    
    # Check if files exist
    missing_files = [f for f, path in files.items() if not os.path.exists(path)]
    if missing_files:
        print(f"Error: The following required files are missing: {', '.join(missing_files)}")
        return None, None, None
    
    # Load dataframes
    try:
        tasks_df = pd.read_excel(files["tasks"])
        contractors_df = pd.read_excel(files["contractors"])
        incidents_df = pd.read_excel(files["incidents"])
        
        # Check for required columns
        req_columns = {
            "tasks": ["Task ID", "Incident ID", "Contractor ID", "Assigned At", "Status", "Status Updated At"],
            "contractors": ["contractor_id", "name", "rating"],
            "incidents": ["Incident ID", "Severity"]
        }
        
        for df_name, df in [("tasks", tasks_df), ("contractors", contractors_df), ("incidents", incidents_df)]:
            missing = [col for col in req_columns[df_name] if col not in df.columns]
            if missing:
                print(f"Error: {df_name}.xlsx is missing required columns: {', '.join(missing)}")
                return None, None, None
        
        return tasks_df, contractors_df, incidents_df
    
    except Exception as e:
        print(f"Error loading data: {e}")
        return None, None, None

def calculate_response_time(row):
    """Calculate response time from Assigned At to Status Updated At for Completed tasks."""
    if row["Status"] != "Completed":
        return np.nan
    
    try:
        # Parse datetime strings to datetime objects
        assigned_at = pd.to_datetime(row["Assigned At"])
        completed_at = pd.to_datetime(row["Status Updated At"])
        
        # Calculate time difference in hours
        response_time = (completed_at - assigned_at).total_seconds() / 3600
        return response_time
    except:
        return np.nan

def is_on_time(row, incidents_df):
    """Check if task was completed within SLA based on incident priority."""
    if row["Status"] != "Completed":
        return np.nan
    
    try:
        # Get incident priority
        incident_id = row["Incident ID"]
        incident = incidents_df[incidents_df["ID"] == incident_id]
        if incident.empty:
            return np.nan
        
        priority = incident["Priority"].iloc[0]
        
        # Define SLA hours by priority
        sla_hours = {
            "Critical": 4,
            "High": 8,
            "Medium": 24,
            "Low": 48
        }
        
        # Default SLA if priority not in our mapping
        default_sla = 24
        allowed_hours = sla_hours.get(priority, default_sla)
        
        # Calculate response time
        assigned_at = pd.to_datetime(row["Assigned At"])
        completed_at = pd.to_datetime(row["Status Updated At"])
        response_time = (completed_at - assigned_at).total_seconds() / 3600
        
        # Check if response time is within SLA
        return response_time <= allowed_hours
    except:
        return np.nan

def generate_performance_report():
    """Generate contractor performance report."""
    # Load data
    tasks_df, contractors_df, incidents_df = load_data()
    if tasks_df is None or contractors_df is None or incidents_df is None:
        return False
    
    # Skip if no completed tasks
    if "Completed" not in tasks_df["Status"].values:
        print("No completed tasks found. Cannot generate performance report.")
        return False
    
    # Calculate response times for completed tasks
    tasks_df["Response Time (Hours)"] = tasks_df.apply(calculate_response_time, axis=1)
    
    # Calculate on-time status
    tasks_df["On Time"] = tasks_df.apply(lambda row: is_on_time(row, incidents_df), axis=1)
    
    # Merge tasks with contractors
    performance_df = pd.merge(
        tasks_df,
        contractors_df,
        left_on="Contractor ID",
        right_on="contractor_id",
        how="left"
    )
    
    # Calculate performance metrics by contractor
    contractor_metrics = []
    
    for contractor_id in performance_df["contractor_id"].unique():
        if pd.isna(contractor_id):
            continue
            
        contractor_tasks = performance_df[performance_df["contractor_id"] == contractor_id]
        completed_tasks = contractor_tasks[contractor_tasks["Status"] == "Completed"]
        
        # Skip if no completed tasks
        if completed_tasks.empty:
            avg_response_time = 0
            on_time_rate = 0
        else:
            avg_response_time = completed_tasks["Response Time (Hours)"].mean()
            on_time_count = completed_tasks["On Time"].sum()
            on_time_rate = on_time_count / len(completed_tasks) * 100 if len(completed_tasks) > 0 else 0
        
        # Get contractor name and rating
        contractor_name = contractor_tasks["name"].iloc[0]
        contractor_rating = contractor_tasks["rating"].iloc[0]
        
        # Count tasks by status
        total_tasks = len(contractor_tasks)
        assigned_tasks = len(contractor_tasks[contractor_tasks["Status"] == "Assigned"])
        in_progress_tasks = len(contractor_tasks[contractor_tasks["Status"] == "In Progress"])
        completed_tasks_count = len(completed_tasks)
        
        contractor_metrics.append({
            "Contractor ID": contractor_id,
            "Contractor Name": contractor_name,
            "Rating": contractor_rating,
            "Total Tasks": total_tasks,
            "Assigned Tasks": assigned_tasks,
            "In Progress Tasks": in_progress_tasks,
            "Completed Tasks": completed_tasks_count,
            "Avg Response Time (Hours)": avg_response_time,
            "On-time Rate (%)": on_time_rate
        })
    
    # Create DataFrame from metrics
    performance_summary = pd.DataFrame(contractor_metrics)
    
    # Calculate overall statistics
    overall_stats = {
        "Contractor ID": "OVERALL",
        "Contractor Name": "System Average",
        "Rating": performance_summary["Rating"].mean(),
        "Total Tasks": performance_summary["Total Tasks"].sum(),
        "Assigned Tasks": performance_summary["Assigned Tasks"].sum(),
        "In Progress Tasks": performance_summary["In Progress Tasks"].sum(),
        "Completed Tasks": performance_summary["Completed Tasks"].sum(),
        "Avg Response Time (Hours)": performance_summary["Avg Response Time (Hours)"].mean(),
        "On-time Rate (%)": performance_summary["On-time Rate (%)"].mean()
    }
    
    # Add overall stats to summary
    performance_summary = performance_summary._append(overall_stats, ignore_index=True)
    
    # Sort by performance metrics (on-time rate, then response time)
    performance_summary = performance_summary.sort_values(
        by=["On-time Rate (%)", "Avg Response Time (Hours)"],
        ascending=[False, True]
    ).reset_index(drop=True)
    
    # Create Excel report
    report_file = "data/contractor_performance.xlsx"
    
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(report_file), exist_ok=True)
    
    # Create Excel writer
    with pd.ExcelWriter(report_file, engine='openpyxl') as writer:
        # Write summary to Excel
        performance_summary.to_excel(writer, sheet_name="Performance Summary", index=False)
        
        # Write task details
        tasks_with_details = performance_df[["Task ID", "Incident ID", "contractor_id", "name", "Assigned At", 
                                           "Status", "Status Updated At", "Response Time (Hours)", "On Time"]]
        tasks_with_details.to_excel(writer, sheet_name="Task Details", index=False)
        
        # Write monthly trend data (assuming Assigned At has date information)
        try:
            performance_df["Month"] = pd.to_datetime(performance_df["Assigned At"]).dt.strftime('%Y-%m')
            monthly_trend = performance_df.groupby(["Month", "contractor_id", "name"]).agg({
                "Task ID": "count",
                "On Time": lambda x: x.mean() * 100,
                "Response Time (Hours)": "mean"
            }).reset_index()
            monthly_trend.columns = ["Month", "Contractor ID", "Contractor Name", 
                                   "Tasks", "On-time Rate (%)", "Avg Response Time (Hours)"]
            monthly_trend.to_excel(writer, sheet_name="Monthly Trends", index=False)
        except:
            # If date parsing fails, skip monthly trends
            print("Warning: Could not generate monthly trends. Check date formats.")
            
    # Apply Excel formatting
    apply_excel_formatting(report_file)
    
    print(f"\nContractor performance report generated: {report_file}")
    return True

def apply_excel_formatting(report_file):
    """Apply conditional formatting and styling to Excel report."""
    try:
        # Load workbook
        wb = load_workbook(report_file)
        
        # Format Performance Summary sheet
        if "Performance Summary" in wb.sheetnames:
            ws = wb["Performance Summary"]
            
            # Define styles
            header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Apply header styles
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                
            # Set column widths
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[chr(64 + col)].width = 18
                
            # Apply conditional formatting for On-time Rate
            on_time_col = None
            response_time_col = None
            
            # Find column indices
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "On-time Rate (%)":
                    on_time_col = col
                elif ws.cell(row=1, column=col).value == "Avg Response Time (Hours)":
                    response_time_col = col
            
            # Apply color scale to on-time rate
            if on_time_col:
                col_letter = chr(64 + on_time_col)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    ColorScaleRule(
                        start_type='num', start_value=0, start_color='F8696B',
                        mid_type='num', mid_value=50, mid_color='FFEB84',
                        end_type='num', end_value=100, end_color='63BE7B'
                    )
                )
            
            # Apply color scale to response time (lower is better)
            if response_time_col:
                col_letter = chr(64 + response_time_col)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    ColorScaleRule(
                        start_type='num', start_value=0, start_color='63BE7B',
                        mid_type='num', mid_value=24, mid_color='FFEB84',
                        end_type='num', end_value=48, end_color='F8696B'
                    )
                )
            
        # Format Task Details sheet
        if "Task Details" in wb.sheetnames:
            ws = wb["Task Details"]
            
            # Apply header styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                
            # Set column widths
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[chr(64 + col)].width = 18
                
            # Find On Time column
            on_time_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "On Time":
                    on_time_col = col
                    break
                    
            # Apply conditional formatting to On Time column
            if on_time_col:
                col_letter = chr(64 + on_time_col)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(
                        operator='equal',
                        formula=['TRUE'],
                        stopIfTrue=True,
                        fill=PatternFill(start_color='63BE7B', end_color='63BE7B', fill_type='solid')
                    )
                )
                
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(
                        operator='equal',
                        formula=['FALSE'],
                        stopIfTrue=True,
                        fill=PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid')
                    )
                )
                
        # Format Monthly Trends sheet
        if "Monthly Trends" in wb.sheetnames:
            ws = wb["Monthly Trends"]
            
            # Apply header styles
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                
            # Set column widths
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save the workbook
        wb.save(report_file)
        
    except Exception as e:
        print(f"Warning: Could not apply formatting to Excel report: {e}")

def main():
    parser = argparse.ArgumentParser(description="Generate contractor performance report in CityInfraXLS")
    parser.add_argument("--output", help="Output file path", default="data/contractor_performance.xlsx")
    args = parser.parse_args()
    
    print("Generating contractor performance report...")
    success = generate_performance_report()
    
    if not success:
        print("Failed to generate performance report.")
        sys.exit(1)
    
    print("Performance report generated successfully.")

if __name__ == "__main__":
    main()