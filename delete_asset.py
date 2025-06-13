# delete_asset.py

import os
import sys
import argparse
import datetime
from pathlib import Path

# Local imports
from utils.excel_handler import load_workbook, save_workbook

# Constants
ASSETS_PATH = 'data/assets.xlsx'
LOG_PATH = 'data/asset_log.xlsx'
LOG_HEADERS = ['Timestamp', 'Asset ID', 'Asset Type', 'Action', 'Details']

def find_asset(asset_id):
    """
    Find an asset by ID in the assets workbook.
    
    Args:
        asset_id (str): The asset ID to search for
        
    Returns:
        tuple: (sheet_name, row_index, asset_data) if found, (None, None, None) otherwise
    """
    # Check if assets file exists
    if not os.path.exists(ASSETS_PATH):
        print(f"Error: Assets file not found at {ASSETS_PATH}")
        return None, None, None
    
    # Load the workbook
    wb = load_workbook(ASSETS_PATH)
    
    # Search through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Find the ID column (assumed to be the first column)
        header_row = [cell.value for cell in ws[1]]
        if 'ID' not in header_row:
            continue
        
        id_col_idx = header_row.index('ID') + 1  # Convert to 1-based index
        
        # Search for the asset ID
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            cell_value = ws.cell(row=row_idx, column=id_col_idx).value
            if str(cell_value) == str(asset_id):
                # Found the asset, collect its data
                asset_data = {}
                for col_idx, header in enumerate(header_row, 1):
                    asset_data[header] = ws.cell(row=row_idx, column=col_idx).value
                
                return sheet_name, row_idx, asset_data
    
    # Asset not found
    return None, None, None

def delete_asset(asset_id, confirm=True):
    """
    Delete an asset by ID and log the deletion.
    
    Args:
        asset_id (str): The asset ID to delete
        confirm (bool): Whether to prompt for confirmation
        
    Returns:
        bool: True if deletion was successful, False otherwise
    """
    # Find the asset
    sheet_name, row_idx, asset_data = find_asset(asset_id)
    
    if not sheet_name:
        print(f"Error: Asset with ID '{asset_id}' not found.")
        return False
    
    # Display asset info
    print("\nAsset Found:")
    for key, value in asset_data.items():
        print(f"{key}: {value}")
    
    # Confirm deletion if required
    if confirm:
        confirmation = input("\nAre you sure you want to delete this asset? (y/N): ")
        if confirmation.lower() != 'y':
            print("Deletion cancelled.")
            return False
    
    try:
        # Load the workbook and get the sheet
        wb = load_workbook(ASSETS_PATH)
        ws = wb[sheet_name]
        
        # Remove the row
        ws.delete_rows(row_idx, 1)
        
        # Save the updated workbook
        save_workbook(wb, ASSETS_PATH)
        
        # Log the deletion
        log_deletion(asset_id, sheet_name, asset_data)
        
        print(f"\nSuccess! Asset with ID '{asset_id}' has been deleted.")
        return True
        
    except Exception as e:
        print(f"Error during asset deletion: {str(e)}")
        return False

def log_deletion(asset_id, asset_type, asset_data):
    """
    Log the deletion action to the asset log file.
    
    Args:
        asset_id (str): The deleted asset ID
        asset_type (str): The type of asset (sheet name)
        asset_data (dict): The data of the deleted asset
    """
    # Ensure log file exists
    if not os.path.exists(LOG_PATH):
        # Create the log file with headers
        from openpyxl import Workbook
        log_wb = Workbook()
        log_ws = log_wb.active
        log_ws.title = "Asset Log"
        for col_idx, header in enumerate(LOG_HEADERS, start=1):
            log_ws.cell(row=1, column=col_idx, value=header)
        log_wb.save(LOG_PATH)
    
    # Load the log workbook
    log_wb = load_workbook(LOG_PATH)
    log_ws = log_wb.active
    
    # Create log entry
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    details = f"Asset deleted: {asset_data.get('Location', 'Unknown location')}"
    
    log_entry = [timestamp, asset_id, asset_type, 'DELETE', details]
    
    # Append log entry
    next_row = log_ws.max_row + 1
    for col_idx, value in enumerate(log_entry, start=1):
        log_ws.cell(row=next_row, column=col_idx, value=value)
    
    # Save the updated log
    save_workbook(log_wb, LOG_PATH)
    
    print(f"Deletion logged in {LOG_PATH}")

def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Delete an asset from CityInfraXLS')
    parser.add_argument('asset_id', help='The ID of the asset to delete')
    parser.add_argument('--force', '-f', action='store_true', help='Delete without confirmation')
    
    args = parser.parse_args()
    
    # Delete the asset
    delete_asset(args.asset_id, confirm=not args.force)

if __name__ == "__main__":
    print("=== CityInfraXLS - Asset Deletion Tool ===")
    try:
        main()
    except KeyboardInterrupt:
        print("\nOperation cancelled.")
    except Exception as e:
        print(f"\nAn unexpected error occurred: {str(e)}")