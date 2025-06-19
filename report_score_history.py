# report_score_history.py

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import uuid

def create_score_history_sheet(workbook_path):
    wb = load_workbook(workbook_path)
    if "ScoreHistory" not in wb.sheetnames:
        sheet = wb.create_sheet("ScoreHistory")
        headers = [
            "SnapshotID", "SnapshotTimestamp", "AssetID",
            "Condition", "DaysSinceMaintenance", "IncidentCount",
            "Score", "RiskCategory", "Quarter", "Year"
        ]
        sheet.append(headers)
        for i, h in enumerate(headers):
            col = chr(65 + i)
            sheet.column_dimensions[col].width = max(15, len(h) + 2)
        wb.save(workbook_path)
        return True
    return False

def append_condition_scores_to_history(workbook_path):
    create_score_history_sheet(workbook_path)
    wb = load_workbook(workbook_path)
    src = wb["Condition Scores"]
    tgt = wb["ScoreHistory"]

    snap_id = str(uuid.uuid4())
    snap_ts = datetime.now()
    quarter = f"Q{(snap_ts.month-1)//3 + 1}"
    year = snap_ts.year

    src_hdr = [c.value for c in src[1]]
    req = ["Asset ID","Condition","Days Since Maintenance","Incident Count","Score","Risk Category"]
    idx = {col: src_hdr.index(col) for col in req}

    rows = 0
    for row in src.iter_rows(min_row=2):
        aid = row[idx["Asset ID"]].value
        if not aid:
            continue
        history_row = [
            snap_id,
            snap_ts,
            aid,
            row[idx["Condition"]].value,
            row[idx["Days Since Maintenance"]].value,
            row[idx["Incident Count"]].value,
            row[idx["Score"]].value,
            row[idx["Risk Category"]].value,
            quarter,
            year
        ]
        tgt.append(history_row)
        rows += 1

    wb.save(workbook_path)
    return rows

def get_score_history_trends(workbook_path, asset_id=None, timeframe=None, group_by=None):
    df = pd.read_excel(workbook_path, sheet_name="ScoreHistory", parse_dates=["SnapshotTimestamp"])
    now = datetime.now()
    if asset_id:
        df = df[df["AssetID"] == asset_id]
    if timeframe == "last_30_days":
        df = df[df["SnapshotTimestamp"] >= now - pd.Timedelta(days=30)]
    elif timeframe == "last_quarter":
        q = (now.month-1)//3 + 1
        df = df[(df["Quarter"]==f"Q{q}")&(df["Year"]==now.year)]
    elif timeframe == "last_year":
        df = df[df["Year"] == now.year]
    if group_by == "quarter":
        grp = df.groupby(["Year","Quarter"])
    elif group_by == "year":
        grp = df.groupby("Year")
    elif group_by == "month":
        df["Month"] = df["SnapshotTimestamp"].dt.month
        grp = df.groupby(["Year","Month"])
    elif group_by == "asset":
        grp = df.groupby("AssetID")
    else:
        return df
    return grp.agg({
        "Score":["mean","min","max"],
        "DaysSinceMaintenance":"mean",
        "IncidentCount":"mean",
        "RiskCategory": lambda x: x.mode().iat[0]
    }).reset_index()

def export_score_history_report(workbook_path, output_path, report_type="quarterly"):
    if report_type == "quarterly":
        df = get_score_history_trends(workbook_path, group_by="quarter")
        pivot = pd.pivot_table(
            df, index=["Year","Quarter"],
            values=["Score","IncidentCount","DaysSinceMaintenance"],
            aggfunc={"Score":["mean","count"],"IncidentCount":"sum","DaysSinceMaintenance":"mean"}
        )
        with pd.ExcelWriter(output_path) as writer:
            pivot.to_excel(writer, "Quarterly Trends")
    elif report_type == "asset_trending":
        df = get_score_history_trends(workbook_path, group_by="asset")
        var = df.groupby("AssetID")["Score"].var().nlargest(10).index
        hist = pd.read_excel(workbook_path, sheet_name="ScoreHistory", parse_dates=["SnapshotTimestamp"])
        with pd.ExcelWriter(output_path) as writer:
            for aid in var:
                sub = hist[hist["AssetID"]==aid].sort_values("SnapshotTimestamp")
                sub.to_excel(writer, f"Asset {aid}", index=False)
    return True

def cleanup_score_history(workbook_path, retention_period=None, max_entries=None):
    df = pd.read_excel(workbook_path, sheet_name="ScoreHistory", parse_dates=["SnapshotTimestamp"])
    orig = len(df)
    if retention_period:
        df = df[df["SnapshotTimestamp"] >= datetime.now() - pd.Timedelta(days=retention_period)]
    if max_entries and len(df) > max_entries:
        df = df.sort_values("SnapshotTimestamp", ascending=False).head(max_entries)
    if len(df) < orig:
        wb = load_workbook(workbook_path)
        if "ScoreHistory" in wb.sheetnames:
            wb.remove(wb["ScoreHistory"])
        sht = wb.create_sheet("ScoreHistory")
        sht.append(df.columns.tolist())
        for _,r in df.iterrows():
            sht.append(r.tolist())
        wb.save(workbook_path)
        return orig - len(df)
    return 0
