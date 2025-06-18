# expense_logger.py

"""
log_expense.py - Tool for logging expenses against budget allocations in CityInfraXLS

This script prompts for expense details, validates against existing budget allocations,
assigns a unique expense ID, and records the expense in the expense tracker sheet.
"""

import os
import json
import argparse
import datetime
import uuid
from decimal import Decimal, InvalidOperation
import pandas as pd
from openpyxl import load_workbook
from utils.excel_handler import create_sheets_from_schema, load_workbook, save_workbook

def generate_expense_id():
    """Generate a unique expense ID with EXP prefix."""
    # Take first 8 characters of a UUID and prefix with EXP
    unique_id = str(uuid.uuid4()).replace('-', '')[:8].upper()
    return f"EXP-{unique_id}"

def load_departments():
    """Load list of departments from budget allocations sheet."""
    budget_path = "data/budget_allocations.xlsx"
    if not os.path.exists(budget_path):
        raise FileNotFoundError(f"Budget allocation file not found: {budget_path}")
    
    try:
        # Load budget allocations
        wb = load_workbook(budget_path)
        ws = wb["Allocations"]
        
        # Get column index for department
        headers = [cell.value for cell in ws[1]]
        dept_index = headers.index("department") + 1  # +1 because openpyxl is 1-indexed
        
        # Extract unique departments
        departments = set()
        for row in ws.iter_rows(min_row=2):  # Skip header row
            dept = row[dept_index-1].value  # -1 to convert back to 0-indexed
            if dept:
                departments.add(dept)
        
        return list(departments)
    except Exception as e:
        raise Exception(f"Error loading departments: {str(e)}")

def validate_budget_available(department, amount, category):
    """
    Validate that sufficient budget is available for the expense.
    
    Args:
        department (str): Department name
        amount (float): Expense amount
        category (str): Expense category
    
    Returns:
        tuple: (is_valid, project_id, fiscal_year, remaining_budget)
    """
    budget_path = "data/budget_allocations.xlsx"
    expense_path = "data/expenses.xlsx"
    
    # Load budget allocations
    budget_df = pd.read_excel(budget_path, sheet_name="Allocations")
    
    # Filter by department and category
    dept_budgets = budget_df[(budget_df["department"] == department) & 
                            (budget_df["category"] == category) &
                            (budget_df["status"].isin(["approved", "allocated"]))]
    
    if dept_budgets.empty:
        return False, None, None, 0
    
    # Get latest fiscal year budget for this department/category
    latest_budget = dept_budgets.sort_values("allocation_date", ascending=False).iloc[0]
    project_id = latest_budget["project_id"]
    fiscal_year = latest_budget["fiscal_year"]
    allocated = float(latest_budget["allocated_amount"])
    
    # Load existing expenses if available
    if os.path.exists(expense_path):
        expense_df = pd.read_excel(expense_path, sheet_name="Expenses")
        
        # Filter expenses by project_id
        project_expenses = expense_df[expense_df["project_id"] == project_id]
        
        # Sum up expenses
        total_spent = project_expenses["amount"].sum() if not project_expenses.empty else 0
    else:
        total_spent = 0
    
    # Calculate remaining budget
    remaining = allocated - total_spent
    
    # Check if there's enough budget
    return remaining >= amount, project_id, fiscal_year, remaining

def prompt_for_category():
    """Prompt user for expense category."""
    categories = ["maintenance", "new_construction", "renovation", 
                  "emergency", "planning", "other"]
    
    print("Available categories:")
    for i, category in enumerate(categories, 1):
        print(f"{i}. {category}")
    
    while True:
        choice = input("Select category (number): ").strip()
        try:
            index = int(choice) - 1
            if 0 <= index < len(categories):
                return categories[index]
            else:
                print("Invalid selection. Please try again.")
        except ValueError:
            print("Please enter a number.")

def append_to_expense_sheet(expense_data):
    """
    Append expense to the expense tracker sheet.
    
    Args:
        expense_data (dict): Expense data to append
    """
    expense_path = "data/expenses.xlsx"
    sheet_name = "Expenses"
    
    # Check if file exists
    if not os.path.exists(expense_path):
        # Create schema on the fly for expenses
        expense_schema = {
            "type": "object",
            "properties": {
                "expense_id": {"type": "string"},
                "project_id": {"type": "string"},
                "department": {"type": "string"},
                "amount": {"type": "number"},
                "category": {"type": "string"},
                "description": {"type": "string"},
                "date": {"type": "string", "format": "date"},
                "fiscal_year": {"type": "string"},
                "recorded_by": {"type": "string"},
                "recorded_on": {"type": "string", "format": "date-time"}
            },
            "required": ["expense_id", "project_id", "department", "amount", 
                         "category", "description", "date", "fiscal_year"]
        }
        
        # Write schema to temporary file
        temp_schema_path = "temp_expense_schema.json"
        with open(temp_schema_path, 'w') as f:
            json.dump(expense_schema, f, indent=2)
        
        # Create Excel file with schema
        create_sheets_from_schema(temp_schema_path, expense_path, sheet_name)
        
        # Clean up temporary schema file
        os.remove(temp_schema_path)
    
    # Load workbook
    wb = load_workbook(expense_path)
    ws = wb[sheet_name]
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Create new row
    new_row = []
    for header in headers:
        new_row.append(expense_data.get(header, ""))
    
    # Append row
    ws.append(new_row)
    
    # Save workbook
    save_workbook(wb, expense_path)
    print(f"Expense recorded in {expense_path}, sheet {sheet_name}")

def main():
    parser = argparse.ArgumentParser(description="Log an expense against a budget allocation")
    parser.add_argument("--batch", action="store_true", help="Run in batch mode without prompts")
    args = parser.parse_args()
    
    try:
        # Load departments
        departments = load_departments()
        
        if not departments:
            print("No budget allocations found. Please allocate a budget first.")
            return 1
        
        # Initialize expense data
        expense_data = {
            "expense_id": generate_expense_id(),
            "recorded_on": datetime.datetime.now().isoformat()
        }
        
        # Department
        if not args.batch:
            print("Available departments:")
            for i, dept in enumerate(departments, 1):
                print(f"{i}. {dept}")
            
            while True:
                choice = input("Select department (number): ").strip()
                try:
                    index = int(choice) - 1
                    if 0 <= index < len(departments):
                        expense_data["department"] = departments[index]
                        break
                    else:
                        print("Invalid selection. Please try again.")
                except ValueError:
                    print("Please enter a number.")
        else:
            expense_data["department"] = input().strip()
            if expense_data["department"] not in departments:
                raise ValueError(f"Department not found: {expense_data['department']}")
        
        # Expense date
        while True:
            date_str = input("Expense date (YYYY-MM-DD, leave empty for today): ").strip()
            if not date_str:
                expense_data["date"] = datetime.date.today().isoformat()
                break
            
            try:
                # Validate date format
                expense_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
                expense_data["date"] = expense_date.isoformat()
                break
            except ValueError:
                print("Invalid date format. Please use YYYY-MM-DD.")
        
        # Expense amount
        while True:
            try:
                amount_str = input("Expense amount ($): ").strip()
                amount = float(Decimal(amount_str))
                if amount <= 0:
                    print("Amount must be positive.")
                    continue
                expense_data["amount"] = amount
                break
            except (ValueError, InvalidOperation):
                print("Please enter a valid number.")
        
        # Expense category
        expense_data["category"] = prompt_for_category()
        
        # Validate against budget
        is_valid, project_id, fiscal_year, remaining = validate_budget_available(
            expense_data["department"], 
            expense_data["amount"],
            expense_data["category"]
        )
        
        if not is_valid:
            print(f"Insufficient budget for {expense_data['department']} in category {expense_data['category']}.")
            proceed = input("Do you want to proceed anyway? (y/n): ").strip().lower()
            if proceed != 'y':
                print("Operation canceled.")
                return 1
        else:
            print(f"Budget available: ${remaining:.2f}")
        
        # Add project_id and fiscal_year to expense data
        expense_data["project_id"] = project_id
        expense_data["fiscal_year"] = fiscal_year
        
        # Expense description
        while True:
            description = input("Expense description: ").strip()
            if len(description) < 5:
                print("Description must be at least 5 characters.")
                continue
            expense_data["description"] = description
            break
        
        # Ask for recorded_by
        recorded_by = input("Recorded by (optional): ").strip()
        if recorded_by:
            expense_data["recorded_by"] = recorded_by
        
        # Append to expense sheet
        append_to_expense_sheet(expense_data)
        
        print(f"Expense {expense_data['expense_id']} for ${expense_data['amount']:.2f} "
              f"has been recorded for {expense_data['department']}.")
        
    except KeyboardInterrupt:
        print("\nOperation canceled.")
        return 1
    except Exception as e:
        print(f"Error: {str(e)}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())