# delete_maintenance.py

import os
import sys
import argparse
import pandas as pd
import shutil
from datetime import datetime
import logging
from openpyxl import load_workbook
import importlib.util

# Configure logging
logging.basicConfig(
    filename='data/maintenance_deletion.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def backup_workbook(excel_path):
    """
    Create a backup of the maintenance history workbook.
    
    Args:
        excel_path (str): Path to the Excel workbook
    
    Returns:
        str: Path to the backup file
    """
    # Create backups directory if it doesn't exist
    backup_dir = "data/backups"
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    
    # Generate backup filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"maintenance_history_backup_{timestamp}.xlsx"
    backup_path = os.path.join(backup_dir, backup_filename)
    
    # Make the backup copy
    shutil.copy2(excel_path, backup_path)
    print(f"Backup created at: {backup_path}")
    logging.info(f"Backup created: {backup_path}")
    
    return backup_path

def verify_maintenance_sheet(excel_path):
    """
    Verify that the Maintenance History sheet exists and has the correct structure.
    Attempts to recreate it if missing.
    
    Args:
        excel_path (str): Path to the Excel workbook
        
    Returns:
        bool: True if sheet exists or was recreated, False if critical error
    """
    try:
        # Try to load the workbook and check if the sheet exists
        workbook = load_workbook(excel_path)
        if "Maintenance History" not in workbook.sheetnames:
            print("Warning: 'Maintenance History' sheet not found in workbook.")
            
            # Create backup before attempting repair
            backup_path = backup_workbook(excel_path)
            print(f"Created backup at {backup_path} before attempting repair.")
            
            # Try to import the excel_handler module
            try:
                # Dynamically import the excel_handler module
                spec = importlib.util.spec_from_file_location(
                    "excel_handler", 
                    os.path.join("utils", "excel_handler.py")
                )
                excel_handler = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(excel_handler)
                
                # Check if the function exists
                if hasattr(excel_handler, 'create_maintenance_history_sheet'):
                    print("Attempting to recreate 'Maintenance History' sheet...")
                    
                    # Call the function to recreate the sheet
                    excel_handler.create_maintenance_history_sheet(excel_path)
                    print("Successfully recreated 'Maintenance History' sheet with schema.")
                    logging.info(f"Recreated 'Maintenance History' sheet in {excel_path}")
                    
                    # Return True since the sheet was recreated
                    return True
                else:
                    print("Error: create_maintenance_history_sheet() function not found in excel_handler module.")
                    logging.error("Function create_maintenance_history_sheet() not found")
                    return False
                    
            except Exception as e:
                print(f"Error importing excel_handler module or recreating sheet: {e}")
                logging.error(f"Failed to import excel_handler or recreate sheet: {e}")
                return False
        
        # Sheet exists, now verify it has at least the record_id column
        workbook.close()  # Close before opening with pandas
        
        # Check sheet structure using pandas
        df = pd.read_excel(excel_path, sheet_name="Maintenance History")
        if 'record_id' not in df.columns:
            print("Error: 'Maintenance History' sheet exists but is missing the 'record_id' column.")
            print("Sheet structure appears to be invalid. Please check the file manually.")
            logging.error("Invalid 'Maintenance History' sheet structure - missing record_id column")
            return False
            
        return True
        
    except Exception as e:
        print(f"Error verifying Maintenance History sheet: {e}")
        logging.error(f"Error verifying maintenance sheet: {e}")
        return False

def delete_maintenance_record(record_id, force=False):
    """
    Delete a maintenance record by ID from the Excel workbook.
    
    Args:
        record_id (str): ID of the record to delete
        force (bool): Skip confirmation if True
    
    Returns:
        bool: True if successful, False otherwise
    """
    excel_path = "data/maintenance_history.xlsx"
    
    # Check if file exists
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found.")
        logging.error(f"File not found: {excel_path}")
        return False
    
    # Verify Maintenance History sheet exists and has correct structure
    if not verify_maintenance_sheet(excel_path):
        print("Error: Cannot proceed with deletion due to issues with the Maintenance History sheet.")
        print("The workbook structure has been repaired if possible, but deletion has been aborted.")
        logging.error(f"Deletion of record {record_id} aborted due to sheet structure issues")
        return False
    
    try:
        # Load workbook into pandas dataframe
        df = pd.read_excel(excel_path, sheet_name="Maintenance History")
        
        # Check if record exists
        record_exists = any(str(r_id) == str(record_id) for r_id in df['record_id'])
        if not record_exists:
            print(f"Error: Record ID {record_id} not found in maintenance history.")
            logging.warning(f"Attempted to delete non-existent record ID: {record_id}")
            return False
        
        # Get the record details for logging
        record_details = df[df['record_id'].astype(str) == str(record_id)].iloc[0].to_dict()
        
        # Confirm deletion if not forced
        if not force:
            print("\nRecord details:")
            for key, value in record_details.items():
                print(f"  {key}: {value}")
            
            confirmation = input(f"\nAre you sure you want to delete record {record_id}? (y/N): ")
            if confirmation.lower() not in ['y', 'yes']:
                print("Deletion cancelled.")
                logging.info(f"Deletion of record {record_id} cancelled by user")
                return False
        
        # Create backup before modifying
        backup_path = backup_workbook(excel_path)
        
        # Remove the record
        df_updated = df[df['record_id'].astype(str) != str(record_id)]
        
        # Check if any record was actually removed
        if len(df_updated) == len(df):
            print(f"Warning: No record was deleted. Record ID may not match exactly.")
            logging.warning(f"Failed to delete record {record_id} - no rows were removed")
            return False
        
        # Save the updated dataframe back to Excel while preserving other sheets
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_updated.to_excel(writer, sheet_name="Maintenance History", index=False)
        
        # Log the deletion
        log_message = f"Deleted record ID {record_id}: {record_details}"
        print(f"Successfully deleted record ID {record_id}")
        print(f"Backup created at: {backup_path}")
        logging.info(log_message)
        
        return True
    
    except Exception as e:
        print(f"Error deleting maintenance record: {e}")
        logging.error(f"Error deleting record ID {record_id}: {e}")
        return False

def main():
    """
    Process command-line arguments and call the delete function.
    """
    parser = argparse.ArgumentParser(description="Delete a maintenance record from the history.")
    parser.add_argument("--record-id", required=True, help="ID of the record to delete")
    parser.add_argument("--force", action="store_true", help="Delete without confirmation")
    
    args = parser.parse_args()
    success = delete_maintenance_record(args.record_id, args.force)
    
    # Provide appropriate exit code for scripting purposes
    if not success:
        sys.exit(1)

if __name__ == "__main__":
    main()