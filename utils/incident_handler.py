# utils/incident_handler.py
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

def create_incident_sheet(file_path="data/incidents.xlsx"):
    """
    Creates an incident tracking Excel file with predefined headers and formatting.
    
    The sheet includes conditional formatting to highlight when elapsed time exceeds SLA.
    
    Args:
        file_path: Path where the Excel file should be created
        
    Returns:
        Path to the created Excel file
    """
    # Ensure data directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidents"
    
    # Define headers
    headers = [
        "Incident ID", 
        "Asset ID", 
        "Reporter", 
        "Type", 
        "Severity", 
        "Reported At", 
        "SLA Deadline", 
        "Status",
        "Elapsed Hours"
    ]
    
    # Apply headers and formatting
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        # Set appropriate column width
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(header) + 2)
    
    # Special formatting for date columns
    date_format = 'yyyy-mm-dd hh:mm'
    col_reported = headers.index("Reported At") + 1
    col_sla = headers.index("SLA Deadline") + 1
    
    # Set number format for date columns
    for col in [col_reported, col_sla]:
        column_letter = get_column_letter(col)
        for row in range(2, 1000):  # Apply to many future rows
            ws.cell(row=row, column=col).number_format = date_format
    
    # Add Elapsed Hours formula
    col_elapsed = headers.index("Elapsed Hours") + 1
    for row in range(2, 1000):  # Apply to many future rows
        elapsed_cell = ws.cell(row=row, column=col_elapsed)
        reported_col = get_column_letter(col_reported)
        
        # Formula: (NOW() - Reported_At) * 24 to convert to hours
        elapsed_cell.value = f'=IF({reported_col}{row}="","",((NOW()-{reported_col}{row})*24))'
        elapsed_cell.number_format = '0.00'
    
    # Add conditional formatting for Elapsed Hours > SLA Deadline
    elapsed_col_letter = get_column_letter(col_elapsed)
    sla_col_letter = get_column_letter(col_sla)
    
    # Create rule: If elapsed hours > SLA deadline, apply red fill
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    formula = f'AND({elapsed_col_letter}2>0,{elapsed_col_letter}2>{sla_col_letter}2)'
    
    # Add rule to the worksheet for the Elapsed Hours column
    ws.conditional_formatting.add(
        f'{elapsed_col_letter}2:{elapsed_col_letter}1000',
        FormulaRule(formula=[formula], fill=red_fill)
    )
    
    # Save the workbook
    wb.save(file_path)
    print(f"Incident tracking sheet created at {file_path}")
    
    return file_path

if __name__ == "__main__":
    # Create the incidents sheet when this module is run directly
    create_incident_sheet()