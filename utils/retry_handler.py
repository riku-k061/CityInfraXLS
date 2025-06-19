# utils/retry_handler.py
import time
import random
import logging
from datetime import datetime
from enum import Enum
from typing import Dict, List, Any, Optional, Callable, Tuple, Union

class RetryableError(Exception):
    """Exception class for errors that can be retried."""
    pass

class PermanentError(Exception):
    """Exception class for errors that should not be retried."""
    pass

class ErrorCategory(str, Enum):
    """Categories of errors for geocoding operations."""
    RATE_LIMIT = "RATE_LIMIT"         # API rate limit exceeded
    AUTH_ERROR = "AUTH_ERROR"         # Invalid credentials or authorization
    INVALID_REQUEST = "INVALID_REQUEST"  # Malformed request
    NOT_FOUND = "NOT_FOUND"           # Address couldn't be geocoded
    SERVER_ERROR = "SERVER_ERROR"     # API server error
    NETWORK_ERROR = "NETWORK_ERROR"   # Network connectivity issues
    UNKNOWN_ERROR = "UNKNOWN_ERROR"   # Unclassified errors

class RetryStrategy:
    """
    Implements exponential backoff retry strategy with jitter.
    """
    
    def __init__(self, 
                 max_retries: int = 5,
                 initial_delay: float = 1.0,
                 max_delay: float = 60.0,
                 backoff_factor: float = 2.0,
                 jitter: bool = True):
        """
        Initialize the retry strategy.
        
        Args:
            max_retries: Maximum number of retry attempts
            initial_delay: Initial delay in seconds
            max_delay: Maximum delay in seconds
            backoff_factor: Multiplier for exponential backoff
            jitter: Whether to add randomness to delays to prevent thundering herd
        """
        self.max_retries = max_retries
        self.initial_delay = initial_delay
        self.max_delay = max_delay
        self.backoff_factor = backoff_factor
        self.jitter = jitter
        
    def get_retry_delay(self, attempt: int) -> float:
        """
        Calculate the delay for a specific retry attempt.
        
        Args:
            attempt: The current retry attempt (0-based)
            
        Returns:
            Delay in seconds before next retry
        """
        if attempt <= 0:
            return 0
            
        delay = min(
            self.max_delay,
            self.initial_delay * (self.backoff_factor ** (attempt - 1))
        )
        
        if self.jitter:
            # Add random jitter (±25% of delay)
            delay = random.uniform(delay * 0.75, delay * 1.25)
            
        return delay
        
    def should_retry(self, attempt: int, error: Exception) -> bool:
        """
        Determine if another retry should be attempted.
        
        Args:
            attempt: The current retry attempt (0-based)
            error: The exception that occurred
            
        Returns:
            Whether to retry
        """
        # Don't retry if max attempts exceeded
        if attempt >= self.max_retries:
            return False
            
        # Don't retry permanent errors
        if isinstance(error, PermanentError):
            return False
            
        # Retry on retryable errors or certain exception types
        return (isinstance(error, RetryableError) or
                isinstance(error, (ConnectionError, TimeoutError)))


class GeocodeLogger:
    """
    Logs geocoding operations and errors to Excel worksheet.
    """
    
    def __init__(self, geodata_file: str):
        """
        Initialize the geocode logger.
        
        Args:
            geodata_file: Path to geodata Excel file
        """
        self.geodata_file = geodata_file
        self.logs = []
        self._ensure_log_sheet()
        
    def _ensure_log_sheet(self) -> None:
        """Create GeocodeLog sheet if it doesn't exist."""
        import pandas as pd
        from openpyxl import load_workbook
        
        try:
            # Check if workbook exists
            wb = load_workbook(self.geodata_file)
            
            # Create sheet if it doesn't exist
            if "GeocodeLog" not in wb.sheetnames:
                # Create a new log sheet with headers
                log_df = pd.DataFrame(columns=[
                    "timestamp", "asset_id", "address", "operation", 
                    "attempt", "status", "error_category", "error_message",
                    "retry_delay", "provider"
                ])
                
                with pd.ExcelWriter(self.geodata_file, engine='openpyxl', mode='a') as writer:
                    log_df.to_excel(writer, sheet_name="GeocodeLog", index=False)
                    
        except FileNotFoundError:
            # File doesn't exist yet, will be created later
            pass
        
    def log_attempt(self, 
                  asset_id: str, 
                  address: str, 
                  operation: str,
                  attempt: int,
                  status: str, 
                  provider: str,
                  error_category: Optional[str] = None,
                  error_message: Optional[str] = None,
                  retry_delay: Optional[float] = None) -> None:
        """
        Log a geocoding attempt.
        
        Args:
            asset_id: Asset identifier
            address: Address being geocoded
            operation: Operation type (e.g., 'geocode', 'validate')
            attempt: Attempt number (1-based)
            status: Status of attempt ('success', 'retry', 'failed')
            provider: Geocoding provider used
            error_category: Category of error if applicable
            error_message: Error message if applicable
            retry_delay: Delay before next retry if applicable
        """
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "asset_id": asset_id,
            "address": address,
            "operation": operation,
            "attempt": attempt,
            "status": status,
            "error_category": error_category,
            "error_message": error_message,
            "retry_delay": retry_delay,
            "provider": provider
        }
        
        self.logs.append(log_entry)
        
        # Write to Excel periodically to avoid memory issues with large batches
        if len(self.logs) >= 50:
            self.flush_logs()
    
    def flush_logs(self) -> None:
        """Write accumulated logs to Excel file."""
        if not self.logs:
            return
            
        import pandas as pd
        
        # Convert logs to DataFrame
        log_df = pd.DataFrame(self.logs)
        
        try:
            # Read existing logs if any
            try:
                existing_logs = pd.read_excel(self.geodata_file, sheet_name="GeocodeLog")
                # Combine with new logs
                combined_logs = pd.concat([existing_logs, log_df], ignore_index=True)
            except:
                # No existing logs or sheet
                combined_logs = log_df
                
            # Write back to Excel
            with pd.ExcelWriter(self.geodata_file, engine='openpyxl', mode='a',
                               if_sheet_exists='replace') as writer:
                combined_logs.to_excel(writer, sheet_name="GeocodeLog", index=False)
                
            # Clear log buffer
            self.logs = []
            
        except Exception as e:
            logging.error(f"Failed to write geocode logs to Excel: {str(e)}")
    
    def generate_summary(self) -> Dict[str, Any]:
        """
        Generate a summary of geocoding operations.
        
        Returns:
            Dictionary with summary statistics
        """
        import pandas as pd
        
        try:
            # Flush any remaining logs
            self.flush_logs()
            
            # Read all logs
            logs = pd.read_excel(self.geodata_file, sheet_name="GeocodeLog")
            
            # Calculate summary statistics
            summary = {
                "total_operations": len(logs),
                "successful_operations": len(logs[logs['status'] == 'success']),
                "failed_operations": len(logs[logs['status'] == 'failed']),
                "retry_operations": len(logs[logs['status'] == 'retry']),
                "most_common_errors": logs[logs['error_category'].notna()]['error_category'].value_counts().to_dict(),
                "average_attempts": logs.groupby('asset_id')['attempt'].max().mean(),
                "operations_by_provider": logs['provider'].value_counts().to_dict()
            }
            
            return summary
            
        except Exception as e:
            logging.error(f"Failed to generate geocode summary: {str(e)}")
            return {
                "error": "Failed to generate summary",
                "reason": str(e)
            }


def with_retry(func: Callable, 
              asset_id: str,
              address: str,
              operation: str,
              provider: str,
              logger: GeocodeLogger,
              retry_strategy: Optional[RetryStrategy] = None) -> Any:
    """
    Execute a function with exponential backoff retry logic.
    
    Args:
        func: Function to execute with retries
        asset_id: Asset ID for logging
        address: Address for logging
        operation: Operation name for logging
        provider: Provider name for logging
        logger: GeocodeLogger instance for logging attempts
        retry_strategy: Optional custom retry strategy
        
    Returns:
        Result of the function call
    
    Raises:
        Exception: If all retries fail
    """
    # Use default retry strategy if none provided
    if retry_strategy is None:
        retry_strategy = RetryStrategy()
        
    attempt = 0
    last_error = None
    
    while True:
        attempt += 1
        
        try:
            # Execute the function
            result = func()
            
            # Log success
            logger.log_attempt(
                asset_id=asset_id,
                address=address,
                operation=operation,
                attempt=attempt,
                status='success',
                provider=provider
            )
            
            return result
            
        except Exception as e:
            last_error = e
            
            # Categorize the error
            error_category, error_message = _categorize_error(e)
            
            # Determine if we should retry
            if retry_strategy.should_retry(attempt, e):
                # Calculate delay for next attempt
                delay = retry_strategy.get_retry_delay(attempt)
                
                # Log retry
                logger.log_attempt(
                    asset_id=asset_id,
                    address=address,
                    operation=operation,
                    attempt=attempt,
                    status='retry',
                    provider=provider,
                    error_category=error_category,
                    error_message=error_message,
                    retry_delay=delay
                )
                
                # Wait before retrying
                time.sleep(delay)
                
            else:
                # Log final failure
                logger.log_attempt(
                    asset_id=asset_id,
                    address=address,
                    operation=operation,
                    attempt=attempt,
                    status='failed',
                    provider=provider,
                    error_category=error_category,
                    error_message=error_message
                )
                
                # Re-raise the exception
                raise
    
    # This should never be reached due to the raise above
    raise last_error


def _categorize_error(error: Exception) -> Tuple[str, str]:
    """
    Categorize an error for logging and retry decisions.
    
    Args:
        error: The exception that occurred
        
    Returns:
        Tuple of (error_category, error_message)
    """
    error_str = str(error).lower()
    
    # Rate limit errors
    if any(term in error_str for term in ['rate limit', 'quota', 'too many requests', '429']):
        return ErrorCategory.RATE_LIMIT.value, str(error)
        
    # Authentication errors
    elif any(term in error_str for term in ['auth', 'key', 'credential', 'permission', '401', '403']):
        return ErrorCategory.AUTH_ERROR.value, str(error)
        
    # Invalid request errors
    elif any(term in error_str for term in ['invalid', 'malformed', 'syntax', 'parameter', '400']):
        return ErrorCategory.INVALID_REQUEST.value, str(error)
        
    # Not found errors
    elif any(term in error_str for term in ['not found', 'no result', 'zero result', '404']):
        return ErrorCategory.NOT_FOUND.value, str(error)
        
    # Server errors
    elif any(term in error_str for term in ['server', 'internal', '500', '502', '503', '504']):
        return ErrorCategory.SERVER_ERROR.value, str(error)
        
    # Network errors
    elif any(term in error_str for term in ['network', 'connection', 'timeout', 'connect']):
        return ErrorCategory.NETWORK_ERROR.value, str(error)
        
    # Default to unknown
    else:
        return ErrorCategory.UNKNOWN_ERROR.value, str(error)