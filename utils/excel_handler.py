# utils/excel_handler.py

import os
import json
import logging
import pandas as pd
from openpyxl import Workbook, load_workbook as openpyxl_load
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import InvalidFileException

# Configure basic logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),  # Log to console
        logging.FileHandler('cityinfraxls.log')  # Log to file
    ]
)

logger = logging.getLogger('excel_handler')

def load_workbook(path):
    """
    Load an Excel workbook from the given file path.
    
    Args:
        path (str): Path to the Excel file
        
    Returns:
        openpyxl.Workbook: Loaded workbook object
    """
    logger.info(f"Loading workbook from {path}")
    try:
        wb = openpyxl_load(path)
        return wb
    except Exception as e:
        logger.error(f"Failed to load workbook from {path}: {str(e)}")
        raise

def save_workbook(wb, path):
    """
    Save a workbook to the specified path.
    
    Args:
        wb (openpyxl.Workbook): Workbook to save
        path (str): Path where to save the workbook
        
    Returns:
        None
    """
    logger.info(f"Saving workbook to {path}")
    try:
        wb.save(path)
    except Exception as e:
        logger.error(f"Failed to save workbook to {path}: {str(e)}")
        raise

def init_workbook(path, headers):
    """
    Initialize a new workbook with the specified headers if it doesn't exist.
    
    Args:
        path (str): Path where to create/check the workbook
        headers (list): List of header names for the first row
        
    Returns:
        openpyxl.Workbook: The initialized or existing workbook
    """
    if os.path.exists(path):
        logger.info(f"Workbook already exists at {path}, loading existing file")
        return load_workbook(path)
    
    logger.info(f"Creating new workbook at {path} with headers: {headers}")
    wb = Workbook()
    ws = wb.active
    
    # Add headers to the first row
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Save the new workbook
    save_workbook(wb, path)
    
    return wb

def create_sheets_from_schema(schema_path, output_path, sheet_name=None):
    """
    Create Excel sheets based on a JSON schema.
    
    Args:
        schema_path (str): Path to the JSON schema file
        output_path (str): Path to save the Excel file
        sheet_name (str, optional): Name of the sheet. Defaults to None.
    """
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Load schema
    with open(schema_path, 'r') as f:
        schema = json.load(f)
    
    if not sheet_name:
        # Convert output_path to sheet_name (e.g., data/assets.xlsx -> assets)
        sheet_name = os.path.basename(output_path).split('.')[0]
    
    # Get properties from schema
    properties = schema.get('properties', {})
    
    # Create DataFrame with columns from schema properties
    df = pd.DataFrame(columns=list(properties.keys()))
    
    # Save empty DataFrame to Excel
    df.to_excel(output_path, index=False, sheet_name=sheet_name)
    print(f"Created {output_path} with columns: {', '.join(df.columns)}")

def create_tasks_sheet(output_path="data/tasks.xlsx", sheet_name="tasks"):
    """
    Create a tasks Excel sheet with predefined columns.
    
    Args:
        output_path (str, optional): Path to save the Excel file. Defaults to "data/tasks.xlsx".
        sheet_name (str, optional): Name of the sheet. Defaults to "tasks".
    """
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Define columns for tasks sheet
    columns = ["Task ID", "Incident ID", "Contractor ID", "Assigned At", "Status", "Details"]
    
    # Create DataFrame with defined columns
    df = pd.DataFrame(columns=columns)
    
    # Save empty DataFrame to Excel
    df.to_excel(output_path, index=False, sheet_name=sheet_name)
    print(f"Created {output_path} with columns: {', '.join(columns)}")

def create_maintenance_history_sheet(path='data/maintenance_history.xlsx'):
    """
    Create a new maintenance history Excel sheet with headers based on the maintenance schema.
    
    Args:
        path (str): Path to the Excel file to create or update
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Load the maintenance schema to get the headers
        with open('maintenance_schema.json', 'r') as schema_file:
            schema = json.load(schema_file)
        
        # Extract the properties to use as headers
        headers = list(schema['properties'].keys())
        
        # Create a new DataFrame with the headers but no data
        df = pd.DataFrame(columns=headers)
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(path), exist_ok=True)
        
        # Create a new workbook
        wb = Workbook()
        # Get the active worksheet and rename it
        ws = wb.active
        ws.title = "Maintenance History"
        
        # Add headers to the sheet
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        
        # Save the workbook
        wb.save(path)
        print(f"Successfully created maintenance history sheet at {path}")
        return True
        
    except Exception as e:
        print(f"Error creating maintenance history sheet: {e}")
        return False
    
def create_complaint_sheet(path):
    """
    Creates a complaint sheet at the specified path if it doesn't exist already.
    If the file exists but doesn't have a Complaints sheet, adds the sheet.
    Ensures the sheet has the correct headers based on the complaint schema.
    
    Args:
        path (str): Path to the Excel file
    
    Returns:
        bool: True if created or already exists with correct structure, False otherwise
    """
    # Load complaint schema
    with open('complaint_schema.json', 'r') as schema_file:
        schema = json.load(schema_file)
    
    # Extract headers from schema properties
    headers = list(schema['properties'].keys())
    
    # If file doesn't exist, create new file with headers
    if not os.path.exists(path):
        df = pd.DataFrame(columns=headers)
        df.to_excel(path, sheet_name='Complaints', index=False)
        return True
    
    try:
        # Check if Complaints sheet exists
        wb = load_workbook(path)
        if 'Complaints' not in wb.sheetnames:
            # If file exists but Complaints sheet doesn't, create it
            with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
                df = pd.DataFrame(columns=headers)
                df.to_excel(writer, sheet_name='Complaints', index=False)
        else:
            # If sheet exists, verify headers
            df = pd.read_excel(path, sheet_name='Complaints')
            existing_headers = df.columns.tolist()
            
            # Check if headers match
            if set(existing_headers) != set(headers):
                # Create a backup
                backup_path = path.replace('.xlsx', '_backup.xlsx')
                wb.save(backup_path)
                
                # Recreate with correct headers, preserving data for matching columns
                with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
                    new_df = pd.DataFrame(columns=headers)
                    
                    # Copy data for columns that exist in both
                    for col in set(existing_headers).intersection(set(headers)):
                        if len(df) > 0:
                            new_df[col] = df[col]
                            
                    # Delete the original sheet
                    wb = load_workbook(path)
                    if 'Complaints' in wb.sheetnames:
                        wb.remove(wb['Complaints'])
                        wb.save(path)
                    
                    # Write the corrected sheet
                    new_df.to_excel(writer, sheet_name='Complaints', index=False)
        
        return True
    
    except (InvalidFileException, Exception) as e:
        print(f"Error creating complaint sheet: {str(e)}")
        return False