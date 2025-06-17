# delete_incident.py
"""
Script to delete an incident from the CityInfraXLS system.
Removes the matching incident row from incidents.xlsx based on incident ID.
"""

import os
import sys
import logging
import argparse
import datetime
import openpyxl
from pathlib import Path
from typing import Optional, Tuple

# Add project root to path to enable imports from other modules
sys.path.insert(0, str(Path(__file__).resolve().parent))

# Configure logging
logging.basicConfig(
    filename='cityinfraxls.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('CityInfraXLS')

def find_incident(file_path: str, incident_id: str) -> Tuple[bool, Optional[dict], Optional[int]]:
    """
    Finds an incident in the Excel file by ID.
    
    Args:
        file_path: Path to the incidents Excel file
        incident_id: The incident ID to search for
    
    Returns:
        Tuple containing:
        - Boolean indicating if the incident was found
        - Dict with incident details if found, None otherwise
        - Row index if found, None otherwise
    """
    # Import here to avoid circular imports
    from utils.incident_handler import create_incident_sheet
    
    # Check if file exists, create it if not
    if not os.path.exists(file_path):
        logger.info(f"Incidents file not found, initializing: {file_path}")
        create_incident_sheet(file_path)
        print(f"No incidents found: Incidents tracking file was just initialized.")
        return False, None, None
    
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Check if the Incidents sheet exists
        try:
            ws = wb["Incidents"]
        except KeyError:
            # Sheet doesn't exist, create it
            logger.info(f"Incidents sheet not found in {file_path}, initializing structure")
            create_incident_sheet(file_path)
            print(f"No incidents found: Incidents tracking sheet was just initialized.")
            return False, None, None
            
        # Get column headers
        headers = [cell.value for cell in ws[1]]
        
        # Search for the incident ID in the first column
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == incident_id:
                # Found the incident - create a dict of incident details
                incident_details = {
                    headers[i]: cell.value 
                    for i, cell in enumerate(row) 
                    if i < len(headers)
                }
                return True, incident_details, row_idx
        
        # Incident not found
        return False, None, None
        
    except Exception as e:
        logger.error(f"Error searching for incident: {e}")
        print(f"Error searching for incident: {e}")
        sys.exit(1)

def delete_incident(file_path: str, row_index: int) -> None:
    """
    Deletes an incident row from the Excel file.
    
    Args:
        file_path: Path to the incidents Excel file
        row_index: The row index to delete
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Incidents"]
        
        # Delete the row
        ws.delete_rows(row_index, 1)
        
        # Save the workbook with error handling
        try:
            wb.save(file_path)
        except PermissionError:
            logger.error(f"Permission denied when saving to {file_path}. File may be open in another program.")
            print(f"Error: Could not save changes. Please ensure the incidents file is not open in another program.")
            sys.exit(1)
        except Exception as e:
            logger.error(f"Failed to save changes to Excel file: {e}")
            print(f"Error: Could not save changes: {str(e)}")
            sys.exit(1)
            
    except Exception as e:
        logger.error(f"Error deleting incident: {e}")
        print(f"Error deleting incident: {e}")
        sys.exit(1)

def main():
    """Main function for incident deletion."""
    parser = argparse.ArgumentParser(description="Delete an incident from CityInfraXLS")
    parser.add_argument("--id", required=True, help="The incident ID to delete")
    parser.add_argument("--force", action="store_true", help="Skip confirmation prompt")
    args = parser.parse_args()
    
    incident_file = 'data/incidents.xlsx'
    
    # Search for the incident
    found, incident_details, row_index = find_incident(incident_file, args.id)
    
    if not found:
        if incident_details is None and row_index is None:
            # This is the case where we just initialized the sheet
            logger.info(f"Delete operation terminated - no incidents exist yet")
            sys.exit(0)
        else:
            # This is the case where the sheet exists but the ID wasn't found
            print(f"Error: No incident found with ID: {args.id}")
            sys.exit(1)
    
    # Display incident details and confirm deletion
    print("\n=== Incident Details ===")
    for key, value in incident_details.items():
        # Format datetime objects for better readability
        if isinstance(value, datetime.datetime):
            value = value.strftime('%Y-%m-%d %H:%M:%S')
        print(f"{key}: {value}")
    
    # Confirm deletion unless --force flag is used
    if not args.force:
        confirm = input("\nAre you sure you want to delete this incident? (y/n): ").strip().lower()
        if confirm != 'y':
            print("Deletion cancelled.")
            sys.exit(0)
    
    # Delete the incident
    delete_incident(incident_file, row_index)
    
    # Log the deletion
    logger.info(f"Incident deleted - ID: {args.id}, Details: {incident_details}")
    
    # Print success message
    print(f"\nSuccess: Incident with ID {args.id} has been deleted.")

if __name__ == "__main__":
    main()