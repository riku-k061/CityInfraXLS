import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta

def predict_maintenance():
    """
    Predicts future maintenance dates based on historical maintenance intervals.
    Creates a new Excel sheet with predictions and a visualization chart.
    """
    print("Starting maintenance prediction analysis...")
    
    # Path to the Excel file
    excel_path = "data/maintenance_history.xlsx"
    
    # Check if file exists
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found.")
        return False
    
    try:
        # Load the maintenance history data
        df = pd.read_excel(excel_path, sheet_name="Maintenance History")
        
        # Ensure date column is datetime type
        df['date'] = pd.to_datetime(df['date'])
        
        # Calculate average maintenance intervals per asset
        intervals_by_asset = {}
        last_date_by_asset = {}
        action_by_asset = {}
        
        # Group by asset_id
        grouped = df.sort_values('date').groupby('asset_id')
        
        for asset_id, group in grouped:
            # Calculate intervals between consecutive maintenance dates
            dates = group['date'].sort_values().reset_index(drop=True)
            
            # Get the last action taken for this asset
            last_action = group.loc[group['date'].idxmax(), 'action_taken'] if not group.empty else "Unknown"
            action_by_asset[asset_id] = last_action
            
            if len(dates) >= 2:
                # Calculate differences between consecutive dates in days
                intervals = []
                for i in range(1, len(dates)):
                    interval = (dates[i] - dates[i-1]).days
                    intervals.append(interval)
                
                # Store average interval and last maintenance date
                intervals_by_asset[asset_id] = np.mean(intervals)
                last_date_by_asset[asset_id] = dates.iloc[-1]
            else:
                # Default interval of 90 days if only one maintenance record
                intervals_by_asset[asset_id] = 90
                last_date_by_asset[asset_id] = dates.iloc[0]
        
        # Create prediction dataframe
        predictions = []
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        for asset_id in intervals_by_asset:
            avg_interval = intervals_by_asset[asset_id]
            last_date = last_date_by_asset[asset_id]
            next_due_date = last_date + timedelta(days=round(avg_interval))
            days_until_due = (next_due_date - today).days
            
            # Determine status based on days until due
            if days_until_due < 0:
                status = "Overdue"
            elif days_until_due <= 7:
                status = "Due Soon"
            elif days_until_due <= 30:
                status = "Upcoming"
            else:
                status = "OK"
            
            predictions.append({
                'Asset ID': asset_id,
                'Last Service Date': last_date,
                'Last Action': action_by_asset.get(asset_id, "Unknown"),
                'Average Interval (days)': round(avg_interval),
                'Next Due Date': next_due_date,
                'Days Until Due': days_until_due,
                'Status': status
            })
        
        predictions_df = pd.DataFrame(predictions)
        
        # Sort by days until due (prioritizing most urgent first)
        predictions_df = predictions_df.sort_values('Days Until Due')
        
        # Save to Excel with chart
        create_prediction_sheet(excel_path, predictions_df, intervals_by_asset)
        
        print(f"Maintenance predictions have been calculated and saved to {excel_path}.")
        return True
        
    except Exception as e:
        print(f"Error analyzing maintenance data: {e}")
        return False

def create_prediction_sheet(excel_path, predictions_df, intervals_by_asset):
    """
    Creates or updates the Predictions sheet in the Excel file and adds a column chart
    showing maintenance intervals per asset.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_path)
    
    # Create or get the Predictions sheet
    if "Predictions" in workbook.sheetnames:
        # Remove the existing sheet to recreate it
        workbook.remove(workbook["Predictions"])
    
    predictions_sheet = workbook.create_sheet("Predictions")
    
    # Add title
    predictions_sheet['A1'] = "MAINTENANCE PREDICTIONS"
    predictions_sheet['A1'].font = Font(size=16, bold=True)
    predictions_sheet.merge_cells('A1:G1')
    predictions_sheet['A1'].alignment = Alignment(horizontal='center')
    
    # Add current date
    predictions_sheet['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    predictions_sheet['A2'].font = Font(italic=True)
    predictions_sheet.merge_cells('A2:G2')
    
    # Add data from the dataframe to the sheet starting from row 4
    start_row = 4
    for i, r in enumerate(dataframe_to_rows(predictions_df, index=False, header=True)):
        predictions_sheet.append(r)
        
        # Add status-based conditional formatting
        if i > 0:  # Skip header row
            row_num = start_row + i
            status_cell = predictions_sheet[f'G{row_num}']
            days_until_due_cell = predictions_sheet[f'F{row_num}']
            
            # Format based on status
            if status_cell.value == "Overdue":
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                days_until_due_cell.font = Font(color="FF0000", bold=True)
            elif status_cell.value == "Due Soon":
                fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
                days_until_due_cell.font = Font(color="FFA500", bold=True)
            elif status_cell.value == "Upcoming":
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            else:
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            
            status_cell.fill = fill
    
    # Apply formatting
    for column in range(1, predictions_df.shape[1] + 1):
        col_letter = openpyxl.utils.get_column_letter(column)
        predictions_sheet.column_dimensions[col_letter].width = 20
    
    # Apply header formatting
    header_row = start_row
    for col in range(1, predictions_df.shape[1] + 1):
        cell = predictions_sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        
        # Add borders to header
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # Get the row count for placing the chart
    data_end_row = start_row + len(predictions_df)
    chart_start_row = data_end_row + 4  # Leave some space after the table
    
    # Create a sorted dictionary for chart data (sorted by asset_id for better readability)
    sorted_intervals = dict(sorted(intervals_by_asset.items()))
    
    # Add chart title
    predictions_sheet.cell(row=chart_start_row-2, column=1).value = "MAINTENANCE INTERVALS BY ASSET"
    predictions_sheet.cell(row=chart_start_row-2, column=1).font = Font(size=14, bold=True)
    
    # Add interval data for charting
    predictions_sheet.cell(row=chart_start_row, column=1).value = "Asset ID"
    predictions_sheet.cell(row=chart_start_row, column=2).value = "Avg. Interval (days)"
    
    for idx, (asset_id, interval) in enumerate(sorted_intervals.items(), 1):
        row = chart_start_row + idx
        predictions_sheet.cell(row=row, column=1).value = asset_id
        predictions_sheet.cell(row=row, column=2).value = round(interval)
    
    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Average Maintenance Intervals by Asset (Days)"
    chart.style = 10
    chart.x_axis.title = "Asset ID"
    chart.y_axis.title = "Interval (Days)"
    
    # Define chart data range
    data = Reference(predictions_sheet, 
                     min_col=2, 
                     min_row=chart_start_row, 
                     max_row=chart_start_row + len(sorted_intervals))
    
    # Define categories
    cats = Reference(predictions_sheet, 
                     min_col=1, 
                     min_row=chart_start_row + 1, 
                     max_row=chart_start_row + len(sorted_intervals))
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Set chart size
    chart.width = 15  # in cm
    chart.height = 10  # in cm
    
    # Add chart to sheet
    predictions_sheet.add_chart(chart, f"A{chart_start_row + len(sorted_intervals) + 2}")
    
    # Save workbook
    workbook.save(excel_path)

if __name__ == "__main__":
    predict_maintenance()