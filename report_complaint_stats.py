# report_complaint_stats.py

"""
CityInfraXLS - Report Complaint Stats Script
Analyzes complaint data and generates department-based statistical reports
"""

import os
import sys
import argparse
import json
import pandas as pd
import numpy as np
from datetime import datetime
import logging
from pathlib import Path
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList

sys.path.append(str(Path(__file__).parent))
from utils.excel_handler import create_complaint_sheet

# Constants
COMPLAINTS_EXCEL = "data/complaints.xlsx"
SCHEMA_PATH = "complaint_schema.json"
LOG_FILE = "cityinfraxls.log"
OUTPUT_DIR = "data/reports"
DATE_FORMAT = "%Y-%m-%d %H:%M:%S"

# Configure logging
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('report_complaint_stats')


def load_schema():
    """Load and validate complaint schema"""
    try:
        with open(SCHEMA_PATH, 'r') as schema_file:
            schema = json.load(schema_file)
            return schema
    except (FileNotFoundError, json.JSONDecodeError) as e:
        error_msg = f"Error loading schema: {str(e)}"
        print(error_msg)
        logger.error(error_msg)
        sys.exit(1)


def load_complaint_data():
    """Load complaint data from Excel file and normalize columns"""
    if not os.path.exists(COMPLAINTS_EXCEL):
        error_msg = f"Error: Complaints file {COMPLAINTS_EXCEL} not found"
        print(error_msg)
        logger.error(error_msg)
        sys.exit(1)
    
    try:
        df = pd.read_excel(COMPLAINTS_EXCEL)
        if df.empty:
            print("No complaint data found in Excel file")
            logger.warning("No complaint data found in Excel file")
            sys.exit(0)
        
        df.columns = [col.lower().replace(' ', '_') for col in df.columns]
        for date_col in ['created_at', 'closed_at']:
            if date_col in df.columns:
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        
        return df
        
    except Exception as e:
        error_msg = f"Error reading complaints Excel: {str(e)}"
        print(error_msg)
        logger.error(error_msg)
        sys.exit(1)


def calculate_resolution_time(row):
    """Calculate resolution time in hours for a complaint"""
    created_at = row['created_at']
    closed_at = row['closed_at']
    
    if pd.notna(closed_at) and pd.notna(created_at):
        # Remove timezone info if present
        if created_at.tzinfo is not None:
            created_at = created_at.tz_localize(None)
        if closed_at.tzinfo is not None:
            closed_at = closed_at.tz_localize(None)
        diff = (closed_at - created_at).total_seconds() / 3600
        return round(diff, 2)
    
    return np.nan


def generate_complaint_stats(df):
    """Generate statistics for complaints grouped by department"""
    df['resolution_time'] = df.apply(calculate_resolution_time, axis=1)
    
    total_complaints = len(df)
    closed_complaints = df['status'].eq('Closed').sum()
    open_complaints = df['status'].eq('Open').sum()
    in_progress_complaints = df['status'].eq('In Progress').sum()
    
    avg_overall_resolution = df['resolution_time'].mean()
    avg_overall_rating = df['rating'].mean() if 'rating' in df.columns else np.nan
    
    if 'department' in df.columns:
        dept_stats = df.groupby('department').agg(
            total_count=('complaint_id', 'count'),
            open_count=('status', lambda x: (x == 'Open').sum()),
            in_progress_count=('status', lambda x: (x == 'In Progress').sum()),
            closed_count=('status', lambda x: (x == 'Closed').sum()),
            avg_resolution_time=('resolution_time', 'mean'),
            avg_rating=('rating', 'mean') if 'rating' in df.columns else ('complaint_id', lambda x: np.nan)
        )
        dept_stats['percent_closed'] = (dept_stats['closed_count'] / dept_stats['total_count'] * 100).round(2)
        dept_stats = dept_stats.sort_values('total_count', ascending=False)
    else:
        dept_stats = pd.DataFrame({
            'total_count': [total_complaints],
            'open_count': [open_complaints],
            'in_progress_count': [in_progress_complaints],
            'closed_count': [closed_complaints],
            'avg_resolution_time': [avg_overall_resolution],
            'avg_rating': [avg_overall_rating],
            'percent_closed': [closed_complaints / total_complaints * 100 if total_complaints > 0 else 0]
        }, index=['All Departments'])
    
    overall_summary = {
        'Total Complaints': total_complaints,
        'Open Complaints': open_complaints,
        'In Progress Complaints': in_progress_complaints,
        'Closed Complaints': closed_complaints,
        'Closure Rate': f"{(closed_complaints / total_complaints * 100):.2f}%" if total_complaints > 0 else "0.00%",
        'Average Resolution Time (hours)': f"{avg_overall_resolution:.2f}" if not np.isnan(avg_overall_resolution) else "N/A",
        'Average Rating': f"{avg_overall_rating:.2f}" if not np.isnan(avg_overall_rating) else "N/A"
    }
    
    return dept_stats, overall_summary, df


def create_styled_excel_report(dept_stats, overall_summary, raw_data, output_path=None):
    """Create a styled Excel report with department statistics"""
    
    # Generate default output path if not provided
    if output_path is None:
        # Create output directory if it doesn't exist
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(OUTPUT_DIR, f"complaint_stats_{timestamp}.xlsx")
    
    # Create a writer object
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Convert department stats for Excel
        dept_stats_excel = dept_stats.copy()
        dept_stats_excel.reset_index(inplace=True)
        dept_stats_excel.columns = [col.replace('_', ' ').title() for col in dept_stats_excel.columns]
        
        # Convert raw data for Excel
        raw_data_excel = raw_data.copy()
        raw_data_excel.columns = [col.replace('_', ' ').title() for col in raw_data_excel.columns]
        
        # Write DataFrames to Excel
        dept_stats_excel.to_excel(writer, sheet_name='Department Stats', index=False)
        
        # Create summary sheet
        summary_df = pd.DataFrame(list(overall_summary.items()), columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Write raw data to Excel
        raw_data_excel.to_excel(writer, sheet_name='Raw Data', index=False)
        
        # Access workbook and worksheets for styling
        workbook = writer.book
        
        # Style Department Stats sheet
        ws_dept = workbook['Department Stats']
        
        # Format header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws_dept[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column width
        for col in ws_dept.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 15)
            ws_dept.column_dimensions[column].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in ws_dept.iter_rows(min_row=1, max_row=ws_dept.max_row):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Skip header row
                    # Center alignment for count columns
                    if any(substr in cell.column_letter for substr in ['B', 'C', 'D', 'E']):
                        cell.alignment = Alignment(horizontal='center')
        
        # Create bar chart for department complaint counts
        chart_sheet = workbook.create_sheet('Charts')
        
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Complaints by Department"
        chart.y_axis.title = "Number of Complaints"
        chart.x_axis.title = "Department"
        
        max_departments = min(10, len(dept_stats_excel))  # Limit to top 10 departments
        
        data = Reference(ws_dept, min_col=2, min_row=1, max_row=max_departments+1, max_col=2)
        categories = Reference(ws_dept, min_col=1, min_row=2, max_row=max_departments+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        chart.height = 15
        chart.width = 20
        
        chart_sheet.add_chart(chart, "A1")
        
        # Create pie chart for status distribution
        pie = PieChart()
        pie.title = "Complaint Status Distribution"
        
        # Create a small table with status counts in the chart sheet
        chart_sheet['A20'] = "Status"
        chart_sheet['B20'] = "Count"
        chart_sheet['A21'] = "Open"
        chart_sheet['B21'] = overall_summary['Open Complaints']
        chart_sheet['A22'] = "In Progress"
        chart_sheet['B22'] = overall_summary['In Progress Complaints'] 
        chart_sheet['A23'] = "Closed"
        chart_sheet['B23'] = overall_summary['Closed Complaints']
        
        data = Reference(chart_sheet, min_col=2, min_row=20, max_row=23)
        categories = Reference(chart_sheet, min_col=1, min_row=21, max_row=23)
        
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(categories)
        
        # Add data labels
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        
        pie.height = 15
        pie.width = 15
        
        chart_sheet.add_chart(pie, "A25")
        
        # Style Summary sheet
        ws_summary = workbook['Summary']
        
        # Format header row
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column width
        for col in ws_summary.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 20)
            ws_summary.column_dimensions[column].width = adjusted_width
        
        # Add borders and styling to all cells
        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Skip header row
                    # Bold font for metric names
                    if cell.column == 1:
                        cell.font = Font(bold=True)
                    # Center alignment for values
                    if cell.column == 2:
                        cell.alignment = Alignment(horizontal='center')
        
        # Style Raw Data sheet
        ws_raw = workbook['Raw Data']
        
        # Format header row
        for cell in ws_raw[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column width (with reasonable maximum)
        for col in ws_raw.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = min(max(max_length, len(str(cell.value))), 50)  # Cap at 50
            adjusted_width = max(max_length + 2, 15)
            ws_raw.column_dimensions[column].width = adjusted_width
    
    print(f"Complaint statistics report created successfully: {output_path}")
    logger.info(f"Complaint statistics report created successfully: {output_path}")
    return output_path


def main():
    """Main function to process command-line arguments"""
    parser = argparse.ArgumentParser(description='Generate complaint statistics report')
    parser.add_argument('--output', help='Output file path (default: data/reports/complaint_stats_TIMESTAMP.xlsx)')
    args = parser.parse_args()
    
    try:
        complaint_data = load_complaint_data()
        dept_stats, overall_summary, enriched_data = generate_complaint_stats(complaint_data)
        
        # Strip timezones from datetime columns before export to Excel
        for col in enriched_data.select_dtypes(include=['datetimetz']).columns:
            enriched_data[col] = enriched_data[col].dt.tz_localize(None)
        
        report_path = create_styled_excel_report(dept_stats, overall_summary, enriched_data, args.output)
        
        print(f"Report generated successfully at: {report_path}")
        return True
        
    except Exception as e:
        error_msg = f"Error generating statistics report: {str(e)}"
        print(error_msg)
        logger.error(error_msg)
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
