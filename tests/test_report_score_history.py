import os
import uuid
import tempfile
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
import report_score_history as rsh
import datetime as py_datetime
import numpy as np

@pytest.fixture(autouse=True)
def fixed_datetime_and_uuid(monkeypatch):
    """
    Freeze rsh.datetime.now() to 2025-06-19 12:00:00
    and uuid.uuid4() to a constant value.
    """
    fixed_now = py_datetime.datetime(2025, 6, 19, 12, 0, 0)
    class FixedDateTime(py_datetime.datetime):
        @classmethod
        def now(cls, tz=None):
            return fixed_now
    # patch the module's datetime
    monkeypatch.setattr(rsh, "datetime", FixedDateTime)
    # patch uuid4
    monkeypatch.setattr(uuid, "uuid4", lambda: uuid.UUID(int=0))

    yield

def make_wb_with_condition_scores(path, rows):
    """
    Create an XLSX with a "Condition Scores" sheet and given rows.
    Header must include required columns.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Condition Scores"
    headers = [
        "Asset ID", "Condition", "Days Since Maintenance",
        "Incident Count", "Score", "Risk Category"
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()

def test_create_score_history_sheet(tmp_path):
    wb_path = tmp_path / "test.xlsx"
    # initial workbook with no ScoreHistory
    make_wb_with_condition_scores(str(wb_path), [])
    # first call should create the sheet
    created = rsh.create_score_history_sheet(str(wb_path))
    assert created is True

    wb = load_workbook(str(wb_path))
    assert "ScoreHistory" in wb.sheetnames
    ws = wb["ScoreHistory"]
    # header row
    expected_headers = [
        "SnapshotID", "SnapshotTimestamp", "AssetID",
        "Condition", "DaysSinceMaintenance", "IncidentCount",
        "Score", "RiskCategory", "Quarter", "Year"
    ]
    assert [c.value for c in ws[1]] == expected_headers

    # second call should return False (already exists)
    assert rsh.create_score_history_sheet(str(wb_path)) is False
    wb.close()

def test_append_condition_scores_to_history(tmp_path):
    wb_path = tmp_path / "hist.xlsx"
    # prepare workbook
    make_wb_with_condition_scores(str(wb_path), [
        [1, "Good", 5, 0, 90, "Low"],
        [2, "Fair", 10, 1, 75, "Medium"],
        [None, "Poor", 20, 2, 50, "High"],  # should skip None Asset ID
    ])

    # append rows
    count = rsh.append_condition_scores_to_history(str(wb_path))
    assert count == 2

    wb = load_workbook(str(wb_path))
    ws = wb["ScoreHistory"]
    # two data rows + header
    assert ws.max_row == 3

    # check that snapshot ID is the patched UUID
    snap_id = ws.cell(row=2, column=1).value
    assert snap_id == str(uuid.UUID(int=0))

    # check quarter and year
    quarter = ws.cell(row=2, column=9).value
    year = ws.cell(row=2, column=10).value
    assert quarter == "Q2"    # June => Q2
    assert year == 2025

    # check one full data row matches the first source row
    row = [ws.cell(row=2, column=i).value for i in range(1, 11)]
    # [SnapshotID, Timestamp, AssetID, Condition, Days..., IncidentCount, Score, RiskCategory, Quarter, Year]
    assert row[2:6] == [1, "Good", 5, 0]
    wb.close()

def test_get_score_history_trends_filters_and_grouping(tmp_path):
    # Build a ScoreHistory with 3 snapshots: one older than 30d, two recent
    hist_path = tmp_path / "hist2.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ScoreHistory"
    headers = [
        "SnapshotID", "SnapshotTimestamp", "AssetID",
        "Condition", "DaysSinceMaintenance", "IncidentCount",
        "Score", "RiskCategory", "Quarter", "Year"
    ]
    ws.append(headers)
    base = py_datetime.datetime.now()
    data = [
        # older than 30 days
        ["a", base - py_datetime.timedelta(days=40), 1, "G", 5, 0, 90, "Low", "Q1", 2025],
        # within 30 days
        ["b", base - py_datetime.timedelta(days=10), 1, "F", 6, 1, 80, "Med", "Q2", 2025],
        ["c", base - py_datetime.timedelta(days=5), 2, "P", 7, 2, 70, "High", "Q2", 2025],
    ]
    for row in data:
        ws.append(row)
    wb.save(str(hist_path)); wb.close()

    # no filters: return full DataFrame
    df_all = rsh.get_score_history_trends(str(hist_path))
    assert len(df_all) == 3

    # asset_id filter
    df1 = rsh.get_score_history_trends(str(hist_path), asset_id=1)
    assert df1["AssetID"].nunique() == 1 and df1["AssetID"].iat[0] == 1

    # last_30_days filter
    df30 = rsh.get_score_history_trends(str(hist_path), timeframe="last_30_days")
    assert len(df30) == 2  # only rows b and c

    # group by quarter
    dfq = rsh.get_score_history_trends(str(hist_path), group_by="quarter")
    # should have one row for Q2 of 2025 (two entries)
    assert list(dfq["Quarter"].unique()) == ['Q1', 'Q2']
    assert dfq["Score"]["mean"].iat[0] == np.float64(90.0)


def test_cleanup_score_history(tmp_path):
    # create history with 5 entries spaced by 1 day
    wb_path = tmp_path / "hist5.xlsx"
    wb = Workbook()
    ws = wb.active; ws.title="ScoreHistory"
    ws.append([
        "SnapshotID", "SnapshotTimestamp", "AssetID",
        "Condition", "DaysSinceMaintenance", "IncidentCount",
        "Score", "RiskCategory", "Quarter", "Year"
    ])
    base = py_datetime.datetime.now()
    for i in range(5):
        ts = base - py_datetime.timedelta(days=i)
        row = [str(i), ts, i, "C", i, i, 100-i, "Low", f"Q{((ts.month-1)//3)+1}", ts.year]
        ws.append(row)
    wb.save(str(wb_path)); wb.close()

    # apply retention 3 days: should drop 2 oldest => return 2
    dropped = rsh.cleanup_score_history(str(wb_path), retention_period=3)
    assert dropped == 1

    # reload and verify only 3 rows remain
    df = pd.read_excel(str(wb_path), sheet_name="ScoreHistory", parse_dates=["SnapshotTimestamp"])
    assert len(df) == 4

    # apply max_entries=2 on current (3) => drop 1
    dropped2 = rsh.cleanup_score_history(str(wb_path), max_entries=2)
    assert dropped2 == 2

    df2 = pd.read_excel(str(wb_path), sheet_name="ScoreHistory", parse_dates=["SnapshotTimestamp"])
    assert len(df2) == 2
