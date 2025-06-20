# tests/test_record_budget.py

import json
import os
import re
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

import record_budget as rb

# --- FIXTURE: SANDBOX CWD & DATA DIR ---

@pytest.fixture(autouse=True)
def tmp_cwd(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()
    return tmp_path

# --- validate_fiscal_year tests ---

def test_validate_fiscal_year_valid():
    assert rb.validate_fiscal_year("2024-2025") == "2024-2025"

def test_validate_fiscal_year_bad_format():
    with pytest.raises(ValueError) as exc:
        rb.validate_fiscal_year("2024/2025")
    assert "Fiscal year must be in format YYYY-YYYY" in str(exc.value)

def test_validate_fiscal_year_bad_sequence():
    with pytest.raises(ValueError) as exc:
        rb.validate_fiscal_year("2024-2027")
    assert "End year must be one year after start year" in str(exc.value)

# --- validate_project_id tests ---

def test_validate_project_id_valid():
    assert rb.validate_project_id("PRJ-000123") == "PRJ-000123"

def test_validate_project_id_invalid():
    with pytest.raises(ValueError) as exc:
        rb.validate_project_id("PJ-123456")
    assert "Project ID must be in format PRJ-XXXXXX" in str(exc.value)

# --- load_schema tests ---

@pytest.fixture
def fake_schema_dir(tmp_path, monkeypatch):
    # Create a fake directory and monkeypatch rb.__file__ to live there
    schema_dir = tmp_path / "schema_dir"
    schema_dir.mkdir()
    fake_file = schema_dir / "record_budget.py"
    fake_file.write_text("")  # contents don't matter
    monkeypatch.setattr(rb, "__file__", str(fake_file))
    return schema_dir

def test_load_schema_missing(fake_schema_dir, capsys):
    # No JSON file → FileNotFoundError (uncaught) or exit?
    # Since load_schema isn't catching FileNotFoundError, it will propagate.
    with pytest.raises(FileNotFoundError):
        rb.load_schema()

def test_load_schema_invalid_json(fake_schema_dir):
    schema_path = fake_schema_dir / "budget_allocation_schema.json"
    schema_path.write_text("{ invalid json ")
    with pytest.raises(json.JSONDecodeError):
        rb.load_schema()

def test_load_schema_valid(fake_schema_dir):
    schema_path = fake_schema_dir / "budget_allocation_schema.json"
    schema_obj = {
        "properties": {
            "category": {"enum": ["A", "B"]},
            "status": {"enum": ["X", "Y"]}
        }
    }
    schema_path.write_text(json.dumps(schema_obj))
    loaded = rb.load_schema()
    assert loaded == schema_obj

# --- prompt_for_department tests ---

def test_prompt_for_department_invalid_then_valid(monkeypatch, capsys):
    inputs = iter(["A", "Dept"])
    monkeypatch.setattr("builtins.input", lambda prompt="": next(inputs))
    result = rb.prompt_for_department()
    out = capsys.readouterr().out
    assert "Department name must be between 2 and 100 characters." in out
    assert result == "Dept"

# --- prompt_for_category tests ---

def test_prompt_for_category_invalid_then_valid(monkeypatch, capsys):
    schema = {"properties": {"category": {"enum": ["CatA", "CatB"]}}}
    # First non-numeric, then out-of-range, then valid "2"
    inputs = iter(["foo", "5", "2"])
    monkeypatch.setattr("builtins.input", lambda prompt="": next(inputs))
    result = rb.prompt_for_category(schema)
    out = capsys.readouterr().out
    assert "Please enter a number." in out
    assert "Invalid selection. Please try again." in out
    assert result == "CatB"

# --- prompt_for_status tests ---

def test_prompt_for_status_invalid_then_valid(monkeypatch, capsys):
    schema = {"properties": {"status": {"enum": ["Open", "Closed"]}}}
    inputs = iter(["0", "3", "1"])
    monkeypatch.setattr("builtins.input", lambda prompt="": next(inputs))
    result = rb.prompt_for_status(schema)
    out = capsys.readouterr().out
    # "0" → index = -1, out-of-range
    assert "Invalid selection. Please try again." in out
    assert result == "Open"

# --- append_to_excel tests ---

@pytest.fixture(autouse=True)
def patch_excel_tools(monkeypatch):
    """
    Stub out create_sheets_from_schema, and wire load/save to openpyxl.
    """
    def stub_create(schema_path, excel_path, sheet_name):
        # Create a fresh workbook with the expected headers
        headers = [
            "allocation_date", "department", "fiscal_year",
            "allocated_amount", "project_id", "category",
            "status", "notes", "approving_authority"
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        wb.save(excel_path)

    monkeypatch.setattr(rb, "create_sheets_from_schema", stub_create)
    monkeypatch.setattr(rb, "load_workbook", lambda path: openpyxl.load_workbook(path))
    monkeypatch.setattr(rb, "save_workbook", lambda wb, path: wb.save(path))

def test_append_creates_file_and_appends(tmp_cwd, capsys):
    record = {
        "allocation_date": date(2025,6,18).isoformat(),
        "department": "Public Works",
        "fiscal_year": "2025-2026",
        "allocated_amount": 12345.67,
        "project_id": "PRJ-123456",
        "category": "CatA",
        "status": "Open",
        "notes": "Test note",
        "approving_authority": "Manager"
    }
    excel_path = "data/budget_allocations.xlsx"
    sheet = "Allocations"

    # File does not exist yet
    assert not os.path.exists(excel_path)
    rb.append_to_excel(excel_path, sheet, record)

    out = capsys.readouterr().out
    assert f"Record appended to {excel_path}, sheet {sheet}" in out
    assert os.path.exists(excel_path)

    # Re-open and check that the row is appended
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet]
    # Header + 1 row → 2 rows
    assert ws.max_row == 2
    values = [cell.value for cell in ws[2]]
    # They should match record in header order
    expected = [
        record["allocation_date"],
        record["department"],
        record["fiscal_year"],
        record["allocated_amount"],
        record["project_id"],
        record["category"],
        record["status"],
        record["notes"],
        record["approving_authority"]
    ]
    assert values == expected

def test_append_existing_missing_sheet(tmp_cwd):
    # Create a workbook with a different sheet
    excel_path = Path("data/budget_allocations.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "WrongSheet"
    wb.save(excel_path)
    record = {"anything": 1}

    with pytest.raises(ValueError) as exc:
        rb.append_to_excel(str(excel_path), "Allocations", record)
    assert "Sheet Allocations not found" in str(exc.value)
