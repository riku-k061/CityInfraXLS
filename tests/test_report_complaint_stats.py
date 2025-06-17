# tests/test_report_complaint_stats.py

import json
import os
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

import report_complaint_stats as rcs

# --- FIXTURE: SANDBOX CWD & DATA DIR ---

@pytest.fixture(autouse=True)
def tmp_cwd(tmp_path, monkeypatch):
    """
    Sandbox each test in tmp_path with a data/ subdirectory.
    """
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()
    return tmp_path

# --- HELPERS ---

def write_schema(obj):
    Path(rcs.SCHEMA_PATH).write_text(json.dumps(obj))

def write_complaints_excel(df: pd.DataFrame):
    path = Path(rcs.COMPLAINTS_EXCEL)
    path.parent.mkdir(exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False)

# --- load_schema tests ---

def test_load_schema_missing(tmp_cwd, capsys):
    # No complaint_schema.json → exit(1)
    with pytest.raises(SystemExit) as exc:
        rcs.load_schema()
    out = capsys.readouterr().out
    assert exc.value.code == 1
    assert "Error loading schema" in out

def test_load_schema_invalid_json(tmp_cwd, capsys):
    # Write malformed JSON
    Path(rcs.SCHEMA_PATH).write_text("{ invalid json ")
    with pytest.raises(SystemExit) as exc:
        rcs.load_schema()
    out = capsys.readouterr().out
    assert exc.value.code == 1
    assert "Error loading schema" in out

def test_load_schema_valid(tmp_cwd):
    # Write a minimal valid schema
    schema = {"properties": {"status": {"enum": ["Open","Closed"]}}}
    write_schema(schema)
    loaded = rcs.load_schema()
    assert isinstance(loaded, dict)
    assert loaded["properties"]["status"]["enum"] == ["Open","Closed"]

# --- load_complaint_data tests ---

def test_load_data_missing_file(tmp_cwd, capsys):
    # No Excel → exit(1)
    with pytest.raises(SystemExit) as exc:
        rcs.load_complaint_data()
    out = capsys.readouterr().out
    assert exc.value.code == 1
    assert "Error: Complaints file" in out

def test_load_data_empty_sheet(tmp_cwd, capsys):
    # Create an empty sheet
    write_complaints_excel(pd.DataFrame())
    with pytest.raises(SystemExit) as exc:
        rcs.load_complaint_data()
    out = capsys.readouterr().out
    assert exc.value.code == 0
    assert "No complaint data found in Excel file" in out

def test_load_data_success(tmp_cwd):
    # Create a sheet with data
    data = {
        "Complaint ID": ["A1"],
        "Status": ["Open"],
        "Created At": ["2025-06-10 08:00:00"],
        "Closed At": ["2025-06-11 10:30:00"],
        "Rating": [4],
        "Department": ["Water"]
    }
    df_in = pd.DataFrame(data)
    write_complaints_excel(df_in)

    df = rcs.load_complaint_data()
    # Columns normalized
    assert "complaint_id" in df.columns
    assert "created_at" in df.columns and df["created_at"].dtype.kind == 'M'
    assert "rating" in df.columns and df["rating"].iloc[0] == 4
    assert df["department"].iloc[0] == "Water"

# --- calculate_resolution_time tests ---

def test_calc_resolution_time_both_dates():
    created = datetime(2025,6,10,8,0)
    closed = datetime(2025,6,11,8,30)
    row = {"created_at": created, "closed_at": closed}
    hours = rcs.calculate_resolution_time(row)
    # 24.5 hours → 24.5 rounded to 2 decimals
    assert hours == 24.5

def test_calc_resolution_time_missing():
    row = {"created_at": datetime.now(), "closed_at": pd.NaT}
    hours = rcs.calculate_resolution_time(row)
    assert np.isnan(hours)

# --- generate_complaint_stats tests ---

@pytest.fixture
def simple_df():
    data = [
        {"complaint_id":"c1","status":"Open","rating":3,"department":"D1",
         "created_at": datetime(2025,6,1,9), "closed_at": pd.NaT},
        {"complaint_id":"c2","status":"Closed","rating":5,"department":"D1",
         "created_at": datetime(2025,6,1,9), "closed_at": datetime(2025,6,2,9)},
        {"complaint_id":"c3","status":"In Progress","rating":4,"department":"D2",
         "created_at": datetime(2025,6,5,10), "closed_at": pd.NaT}
    ]
    return pd.DataFrame(data)

def test_generate_stats(simple_df):
    dept_stats, overall, _ = rcs.generate_complaint_stats(simple_df)
    # D1 has 2 complaints
    assert dept_stats.loc["D1","total_count"] == 2
    # percent_closed for D1 = 1/2*100 = 50.0
    assert dept_stats.loc["D1","percent_closed"] == 50.0
    # overall summary
    assert overall["Total Complaints"] == 3
    assert overall["Closed Complaints"] == 1
    assert overall["Closure Rate"] == "33.33%"

# --- create_styled_excel_report tests ---

def test_create_styled_excel_report_custom_path(tmp_cwd):
    # Prepare dummy stats and raw data
    idx = ["D1","D2"]
    dept_stats = pd.DataFrame({
        "total_count":[2,1],
        "open_count":[1,0],
        "in_progress_count":[0,1],
        "closed_count":[1,0],
        "avg_resolution_time":[24.0, np.nan],
        "avg_rating":[4.0,4.0],
        "percent_closed":[50.0,0.0]
    }, index=idx)
    overall = {
        "Total Complaints": 3,
        "Open Complaints": 1,
        "In Progress Complaints": 1,
        "Closed Complaints": 1,
        "Closure Rate": "33.33%",
        "Average Resolution Time (hours)": "24.00",
        "Average Rating": "4.00"
    }
    raw_data = pd.DataFrame({
        "complaint_id":["c1","c2","c3"],
        "status":["Open","Closed","In Progress"],
        "rating":[3,5,4],
        "department":["D1","D1","D2"],
        "created_at":[datetime(2025,6,1,9)]*3,
        "closed_at":[pd.NaT, datetime(2025,6,2,9), pd.NaT]
    })

    out_path = "my_report.xlsx"
    returned = rcs.create_styled_excel_report(dept_stats, overall, raw_data, output_path=out_path)

    # It should return the same path, and file must exist
    assert returned == out_path
    assert Path(out_path).exists()

    # Inspect workbook
    wb = load_workbook(out_path)
    assert set(wb.sheetnames) >= {"Department Stats","Summary","Raw Data","Charts"}

    # Quick check: Summary sheet's A2 should read "Open Complaints"
    ws = wb["Summary"]
