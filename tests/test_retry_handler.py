# tests/test_retry_handler.py

import os
import random
import time
import pandas as pd
import pytest
from datetime import datetime
from openpyxl import load_workbook, Workbook

import utils.retry_handler as rh


@pytest.fixture(autouse=True)
def no_sleep(monkeypatch):
    # avoid actual sleeping
    monkeypatch.setattr(time, "sleep", lambda s: None)


def test_get_retry_delay_no_jitter():
    strat = rh.RetryStrategy(max_retries=5, initial_delay=1.0, max_delay=10.0,
                             backoff_factor=2.0, jitter=False)
    # attempt 0: no delay
    assert strat.get_retry_delay(0) == 0
    # attempt 1: 1 * 2^(1-1) = 1
    assert strat.get_retry_delay(1) == 1.0
    # attempt 2: 1 * 2^(2-1) = 2
    assert strat.get_retry_delay(2) == 2.0
    # attempt 3: 1 * 2^(3-1) = 4
    assert strat.get_retry_delay(3) == 4.0
    # attempt 10: would exceed max_delay, so capped to 10
    assert strat.get_retry_delay(10) == 10.0

def test_get_retry_delay_with_jitter(monkeypatch):
    # patch random.uniform to a known value
    monkeypatch.setattr(random, "uniform", lambda a, b: (a + b) / 2)
    strat = rh.RetryStrategy(initial_delay=2.0, backoff_factor=2.0, max_delay=100.0, jitter=True)
    # attempt 2: base delay 4, jitter returns midpoint of [3,5] = 4
    d = strat.get_retry_delay(2)
    assert pytest.approx(d, rel=1e-3) == 4.0

def test_should_retry_logic():
    strat = rh.RetryStrategy(max_retries=2)
    # RetryableError should retry when attempt < max
    assert strat.should_retry(0, rh.RetryableError("retry")) is True
    # ConnectionError
    assert strat.should_retry(0, ConnectionError("conn")) is True
    # TimeoutError
    assert strat.should_retry(0, TimeoutError("timeout")) is True
    # PermanentError should not retry
    assert strat.should_retry(0, rh.PermanentError("perm")) is False
    # Exceeded max_retries
    assert strat.should_retry(2, rh.RetryableError("retry")) is False
    # Other exceptions not retryable
    assert strat.should_retry(0, ValueError("fail")) is False

@pytest.mark.parametrize("msg,expected", [
    ("Rate limit exceeded", rh.ErrorCategory.RATE_LIMIT.value),
    ("HTTP 401 Unauthorized", rh.ErrorCategory.AUTH_ERROR.value),
    ("400 Bad request", rh.ErrorCategory.INVALID_REQUEST.value),
    ("Not found", rh.ErrorCategory.NOT_FOUND.value),
    ("Internal server error 500", rh.ErrorCategory.SERVER_ERROR.value),
    ("Network connectivity failed", rh.ErrorCategory.NETWORK_ERROR.value),
    ("Some other error", rh.ErrorCategory.UNKNOWN_ERROR.value),
])
def test_categorize_error(msg, expected):
    cat, full = rh._categorize_error(Exception(msg))
    assert cat == expected
    assert msg in full

@pytest.fixture
def log_file(tmp_path):
    # create an empty workbook so flush_logs can append
    p = tmp_path / "geodata.xlsx"
    Workbook().save(p)
    return str(p)

def test_geocode_logger_ensure_and_flush_and_summary(log_file, tmp_path):
    logger = rh.GeocodeLogger(log_file)
    # GeocodeLog sheet should now exist
    wb = load_workbook(log_file)
    assert "GeocodeLog" in wb.sheetnames
    wb.close()

    # Log a few attempts
    logger.log_attempt("A1", "addr1", "op", 1, "success", "prov")
    logger.log_attempt("A2", "addr2", "op", 1, "retry", "prov", error_category="RATE_LIMIT", error_message="limit", retry_delay=0.5)
    logger.log_attempt("A1", "addr1", "op", 2, "failed", "prov", error_category="UNKNOWN_ERROR", error_message="fail")

    # buffer < 50, so flush explicitly
    logger.flush_logs()

    # Read back the sheet
    df = pd.read_excel(log_file, sheet_name="GeocodeLog")
    # Should have exactly 3 rows
    assert len(df) == 3
    # Check columns
    for col in ["asset_id", "status", "error_category", "retry_delay", "provider"]:
        assert col in df.columns

    # Test summary
    summary = logger.generate_summary()
    assert summary["total_operations"] == 3
    assert summary["successful_operations"] == 1
    assert summary["retry_operations"] == 1
    assert summary["failed_operations"] == 1
    # error categories counted
    assert summary["most_common_errors"].get("RATE_LIMIT", 0) == 1
    assert "prov" in summary["operations_by_provider"]

def test_with_retry_success_and_retries(log_file):
    # function fails twice then succeeds
    calls = {"n":0}
    def flaky():
        calls["n"] += 1
        if calls["n"] < 3:
            raise rh.RetryableError("please retry")
        return "done"

    logger = rh.GeocodeLogger(log_file)
    strat = rh.RetryStrategy(max_retries=5, initial_delay=0.0, jitter=False)
    result = rh.with_retry(
        func=flaky,
        asset_id="X1",
        address="addr",
        operation="geocode",
        provider="test",
        logger=logger,
        retry_strategy=strat
    )
    assert result == "done"
    # flush and inspect log
    logger.flush_logs()
    df = pd.read_excel(log_file, sheet_name="GeocodeLog")
    # attempts: first retry, second retry, third success
    statuses = df.status.tolist()
    assert statuses == ["retry","retry","success"]

def test_with_retry_permanent_error(log_file):
    def always_bad():
        raise rh.PermanentError("cannot retry")

    logger = rh.GeocodeLogger(log_file)
    with pytest.raises(rh.PermanentError):
        rh.with_retry(always_bad, "A", "addr", "op", "prov", logger)
    logger.flush_logs()
    df = pd.read_excel(log_file, sheet_name="GeocodeLog")
    # only one failed log
    assert len(df) == 1
    assert df.iloc[0].status == "failed"

def test_with_retry_exhausted_retries(log_file):
    # always raises retryable
    def bad():
        raise rh.RetryableError("oops")

    logger = rh.GeocodeLogger(log_file)
    strat = rh.RetryStrategy(max_retries=2, initial_delay=0.0, jitter=False)
    with pytest.raises(rh.RetryableError):
        rh.with_retry(bad, "B", "addr", "op", "prov", logger, retry_strategy=strat)
    logger.flush_logs()
    df = pd.read_excel(log_file, sheet_name="GeocodeLog")
    # first attempt logged as retry, second as failed
    statuses = df.status.tolist()
    assert statuses == ["retry", "failed"]
