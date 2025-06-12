# tests/test_incident_handler.py

import os
import pytest
import openpyxl
from utils.incident_handler import create_incident_sheet

@pytest.fixture
def incident_file_path(tmp_path):
    """Fixture to generate a temporary file path"""
    return tmp_path / "data" / "test_incidents.xlsx"

def test_create_incident_sheet_creates_file_and_headers(incident_file_path):
    """Test that the incident sheet is created with expected headers and formatting"""
    # Execute
    result_path = create_incident_sheet(str(incident_file_path))

    # Assert file was created
    assert os.path.exists(result_path)

    # Load workbook
    wb = openpyxl.load_workbook(result_path)
    ws = wb.active

    # Check sheet title
    assert ws.title == "Incidents"

    # Expected headers
    expected_headers = [
        "Incident ID", "Asset ID", "Reporter", "Type", "Severity",
        "Reported At", "SLA Deadline", "Status", "Elapsed Hours"
    ]
    
    # Validate headers
    actual_headers = [ws.cell(row=1, column=i+1).value for i in range(len(expected_headers))]
    assert actual_headers == expected_headers

    # Verify formatting: bold + center alignment
    for i in range(1, len(expected_headers) + 1):
        cell = ws.cell(row=1, column=i)
        assert cell.font.bold
        assert cell.alignment.horizontal == "center"

    # Validate Elapsed Hours formula in second row
    elapsed_formula = ws.cell(row=2, column=len(expected_headers)).value
    assert "NOW()" in str(elapsed_formula)
    assert "*24" in str(elapsed_formula)

    wb.close()