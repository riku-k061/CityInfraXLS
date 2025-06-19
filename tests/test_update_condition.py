import os
import json
import datetime as py_datetime
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

import utils.excel_handler as eh

@pytest.fixture(autouse=True)
def fixed_today_and_strptime(monkeypatch):
    """
    Freeze eh.datetime.datetime.now() to 2025-06-19 and
    patch eh.datetime.strptime so calculate_days_since_maintenance works.
    """
    # 1) Freeze now()
    class FixedDateTime(py_datetime.datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2025, 6, 19)
    monkeypatch.setattr(eh.datetime, 'datetime', FixedDateTime)

    # 2) Inject strptime onto the datetime module
    monkeypatch.setattr(eh.datetime, 'strptime', py_datetime.datetime.strptime, raising=False)
    yield

def create_excel(path, sheet_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()

def test_calculate_days_since_maintenance_various():
    # valid ISO string
    assert eh.calculate_days_since_maintenance("2025-06-01") == 18
    # None => None
    assert eh.calculate_days_since_maintenance(None) is None
    # malformed string => None
    assert eh.calculate_days_since_maintenance("bad-date") is None
    # wrong type => None
    assert eh.calculate_days_since_maintenance(12345) is None

def test_update_condition_scores(tmp_path):
    # 1) assets.xlsx with Road sheet
    assets = tmp_path / "assets.xlsx"
    create_excel(
        str(assets),
        "Road",
        ["ID", "Last Maintenance"],
        [
            [101, "2025-06-10"],
            [102, "2025-06-01"],
            [103, None],
        ]
    )

    # 2) scores.xlsx with Condition Scores
    scores = tmp_path / "scores.xlsx"
    create_excel(
        str(scores),
        "Condition Scores",
        ["Asset ID", "Score"],
        [
            [101, 5],
            [102, 3],
            [103, 4],
        ]
    )

    # Act
    result = eh.update_condition_scores(str(assets), str(scores))
    assert result is True

    # Reload and verify
    wb = load_workbook(str(scores))
    ws = wb["Condition Scores"]
    headers = [c.value for c in ws[1]]
    assert "Days Since Maintenance" in headers
    col_idx = headers.index("Days Since Maintenance") + 1

    # Row 2: 2025-06-10 → 9 days
    assert ws.cell(row=2, column=col_idx).value == 9
    # Row 3: 2025-06-01 → 18 days
    assert ws.cell(row=3, column=col_idx).value == 18
    # Row 4: None → None
    assert ws.cell(row=4, column=col_idx).value is None
    wb.close()

def test_calculate_next_maintenance(tmp_path):
    # 1) Intervals file
    intervals = tmp_path / "intervals.xlsx"
    pd.DataFrame({
        "Asset Type": ["Asphalt", "Concrete"],
        "Maintenance Interval (days)": [30, 60]
    }).to_excel(str(intervals), index=False)

    # 2) Assets file with Road sheet
    assets = tmp_path / "assets2.xlsx"
    create_excel(
        str(assets),
        "Road",
        ["ID", "Surface Type", "Last Maintenance", "Next Maintenance Due"],
        [
            [201, "Asphalt", "2025-06-01", None],
            [202, "Concrete", py_datetime.datetime(2025, 5, 1), None],
            [203, "Unknown", "2025-06-10", None],  # invalid type
            [204, "Asphalt", None, None],          # missing date
        ]
    )

    errs = eh.calculate_next_maintenance(str(assets), str(intervals), error_log=True)
    # returns 0 (error count excluding header)
    assert errs == 0

    wb = load_workbook(str(assets))
    ws = wb["Road"]
    hdrs = [c.value for c in ws[1]]
    nidx = hdrs.index("Next Maintenance Due") + 1

    # 201 → 2025-06-01 + 30d = 2025-07-01
    v1 = ws.cell(row=2, column=nidx).value
    assert getattr(v1, 'date', lambda: v1)() == py_datetime.date(2025, 7, 1)

    # 202 → 2025-05-01 + 60d = 2025-06-30
    v2 = ws.cell(row=3, column=nidx).value
    assert getattr(v2, 'date', lambda: v2)() == py_datetime.date(2025, 6, 30)

    # 203 & 204 remain blank
    assert ws.cell(row=4, column=nidx).value is None
    assert ws.cell(row=5, column=nidx).value is None

    # check ScoringErrors sheet exists and has two entries + header
    assert "ScoringErrors" in wb.sheetnames
    err_ws = wb["ScoringErrors"]
    assert err_ws.max_row == 3  # header + 2 error rows
    wb.close()
