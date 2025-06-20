# tests/test_expense_logger.py

import os
import re
import uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

import expense_logger as el

# --- FIXTURE: SANDBOX CWD & DATA DIR ---

@pytest.fixture(autouse=True)
def tmp_cwd(tmp_path, monkeypatch):
    """
    Run tests in a temp directory with a data/ subfolder.
    """
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()
    return tmp_path

# --- generate_expense_id tests ---

def test_generate_expense_id_format_and_uniqueness():
    ids = {el.generate_expense_id() for _ in range(5)}
    # All must start with EXP- and 8 hex chars
    pat = re.compile(r"^EXP-[0-9A-F]{8}$")
    for eid in ids:
        assert pat.match(eid)
    # They should all be unique
    assert len(ids) == 5

# --- load_departments tests ---

def write_budget_allocations(rows):
    """
    Helper: write data/budget_allocations.xlsx with sheet 'Allocations'
    """
    path = Path("data/budget_allocations.xlsx")
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Allocations", index=False)
    return path

def test_load_departments_file_not_found():
    with pytest.raises(FileNotFoundError):
        el.load_departments()

def test_load_departments_success(tmp_cwd):
    rows = [
        {"project_id":"P1","department":"D1","category":"C","status":"allocated",
         "allocation_date":"2025-01-01","allocated_amount":100},
        {"project_id":"P2","department":"D2","category":"C","status":"approved",
         "allocation_date":"2025-02-01","allocated_amount":200},
        {"project_id":"P3","department":"D1","category":"C","status":"allocated",
         "allocation_date":"2025-03-01","allocated_amount":150},
    ]
    write_budget_allocations(rows)
    depts = el.load_departments()
    assert set(depts) == {"D1", "D2"}

# --- validate_budget_available tests ---

@pytest.fixture
def budget_and_expenses(tmp_cwd):
    """
    Create budget_allocations.xlsx and optionally an expenses.xlsx
    """
    allocs = [
        {"project_id":"PRJ1","department":"DeptA","category":"Cat1","status":"approved",
         "allocation_date":"2025-01-01","allocated_amount":100,"fiscal_year":"2025-2026"},
        {"project_id":"PRJ2","department":"DeptA","category":"Cat1","status":"allocated",
         "allocation_date":"2025-06-01","allocated_amount":50,"fiscal_year":"2025-2026"},
    ]
    write_budget_allocations(allocs)
    yield
    exp_path = Path("data/expenses.xlsx")
    if exp_path.exists():
        exp_path.unlink()

def test_validate_no_budget(tmp_cwd):
    # No budget_allocations.xlsx => FileNotFoundError
    with pytest.raises(FileNotFoundError):
        el.validate_budget_available("DeptX", 10, "Cat1")

def test_validate_sufficient_when_no_expenses(budget_and_expenses):
    ok, pid, fy, rem = el.validate_budget_available("DeptA", 40, "Cat1")
    # Latest budget is PRJ2 with 50
    assert ok is True
    assert pid == "PRJ2"
    assert fy == "2025-2026"
    assert rem == 50

def test_validate_insufficient_when_over(budget_and_expenses):
    ok, pid, fy, rem = el.validate_budget_available("DeptA", 60, "Cat1")
    assert ok is False
    assert pid == "PRJ2"
    assert fy == "2025-2026"
    assert rem == 50

def test_validate_with_existing_expenses(budget_and_expenses):
    # Create an expenses.xlsx for PRJ2 with amount 30
    exp_path = Path("data/expenses.xlsx")
    df = pd.DataFrame([
        {
            "expense_id":"E1","project_id":"PRJ2","department":"DeptA","amount":30,
            "category":"Cat1","description":"x","date":"2025-06-02","fiscal_year":"2025-2026",
            "recorded_by":"U","recorded_on":"2025-06-02T12:00:00"
        }
    ])
    with pd.ExcelWriter(exp_path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Expenses", index=False)

    ok, pid, fy, rem = el.validate_budget_available("DeptA", 25, "Cat1")
    # 50 allocated - 30 spent = 20 remaining
    assert not ok
    assert rem == 20

# --- prompt_for_category tests ---

def test_prompt_for_category_invalid_then_valid(monkeypatch, capsys):
    inputs = iter(["foo", "10", "3"])
    monkeypatch.setattr("builtins.input", lambda prompt="": next(inputs))
    cat = el.prompt_for_category()
    out = capsys.readouterr().out
    assert "Please enter a number." in out
    assert "Invalid selection. Please try again." in out
    # 3 → "renovation"
    assert cat == "renovation"

# --- append_to_expense_sheet tests ---

@pytest.fixture(autouse=True)
def patch_excel_tools(monkeypatch):
    """
    Stub out create_sheets_from_schema, load_workbook, save_workbook to use real openpyxl.
    """
    import expense_logger as module

    def stub_create(schema_path, excel_path, sheet_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        headers = [
            "expense_id","project_id","department","amount","category",
            "description","date","fiscal_year","recorded_by","recorded_on"
        ]
        ws.append(headers)
        wb.save(excel_path)

    monkeypatch.setattr(module, "create_sheets_from_schema", stub_create)
    monkeypatch.setattr(module, "load_workbook", lambda path: openpyxl.load_workbook(path))
    monkeypatch.setattr(module, "save_workbook", lambda wb, path: wb.save(path))

def test_append_creates_and_appends(tmp_cwd, capsys):
    expense = {
        "expense_id": "EXP-ABC12345",
        "project_id": "PRJ1",
        "department": "DeptA",
        "amount": 25.5,
        "category": "Cat1",
        "description": "Test exp",
        "date": date.today().isoformat(),
        "fiscal_year": "2025-2026",
        "recorded_by": "User1",
        "recorded_on": datetime.now().isoformat()
    }
    exp_path = Path("data/expenses.xlsx")
    # No file yet
    assert not exp_path.exists()

    el.append_to_expense_sheet(expense)
    out = capsys.readouterr().out
    assert "Expense recorded in data/expenses.xlsx, sheet Expenses" in out
    assert exp_path.exists()

    wb = openpyxl.load_workbook(exp_path)
    ws = wb["Expenses"]
    # header + one row
    assert ws.max_row == 2
    row_vals = [cell.value for cell in ws[2]]
    expected = [expense[h] for h in [cell.value for cell in ws[1]]]
    assert row_vals == expected

def test_append_existing_missing_sheet(tmp_cwd):
    # Create a workbook with wrong sheet name
    exp_path = Path("data/expenses.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Wrong"
    wb.save(exp_path)

    # Accessing a missing sheet raises KeyError
    with pytest.raises(KeyError):
        el.append_to_expense_sheet({"anything": 1})
