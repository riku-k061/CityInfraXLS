# tests/test_log_expense.py

import json
import os
import re
import shutil
import tempfile
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

import log_expense as le

# --- SANDBOX CWD & DATA DIR ---

@pytest.fixture(autouse=True)
def tmp_env(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()
    yield
    # Cleanup any temp schema file
    for f in tmp_path.glob("temp_expense_schema.json"):
        f.unlink()

# --- generate_expense_id ---

def test_generate_expense_id():
    ids = {le.generate_expense_id() for _ in range(10)}
    pat = re.compile(r"^EXP-[0-9A-F]{8}$")
    assert all(pat.match(e) for e in ids)
    assert len(ids) == 10

# --- load_departments ---

def write_allocations(rows):
    path = Path("data/budget_allocations.xlsx")
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Allocations", index=False)

def test_load_departments_missing():
    with pytest.raises(FileNotFoundError):
        le.load_departments()

def test_load_departments_success():
    rows = [
        {"project_id":"P1","department":"D1","category":"C","status":"allocated","allocation_date":"2025-01-01"},
        {"project_id":"P2","department":"D2","category":"C","status":"approved", "allocation_date":"2025-02-01"}
    ]
    write_allocations(rows)
    depts = le.load_departments()
    assert set(depts) == {"D1","D2"}

# --- get_budget_info ---

@pytest.fixture
def budget_only(tmp_env):
    allocs = [
        {"project_id":"PRJ1","department":"DeptX","category":"Cat1","status":"approved",
         "allocation_date":"2025-01-01","allocated_amount":100,"fiscal_year":"2025-2026"},
        {"project_id":"PRJ2","department":"DeptX","category":"Cat1","status":"allocated",
         "allocation_date":"2025-06-01","allocated_amount":50,"fiscal_year":"2025-2026"},
    ]
    write_allocations(allocs)
    yield

def test_get_budget_info_none(tmp_env):
    # Create a sheet with correct headers but no data rows
    df = pd.DataFrame(columns=["project_id","department","category","status","allocation_date","allocated_amount","fiscal_year"])
    with pd.ExcelWriter("data/budget_allocations.xlsx", engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Allocations", index=False)

    row, pid, fy, alloc, spent, rem = le.get_budget_info("X","Y")
    assert row is None and pid is None and fy is None
    assert alloc == 0 and spent == 0 and rem == 0

def test_get_budget_info_no_expenses(budget_only):
    row, pid, fy, alloc, spent, rem = le.get_budget_info("DeptX","Cat1")
    assert row == 1
    assert pid == "PRJ2"
    assert fy == "2025-2026"
    assert alloc == 50
    assert spent == 0
    assert rem == 50

def test_get_budget_info_with_expenses(budget_only):
    # add an expenses.xlsx
    exp = [
        {"expense_id":"E1","project_id":"PRJ2","department":"DeptX","amount":30,
         "category":"Cat1","description":"d","date":"2025-06-02","fiscal_year":"2025-2026",
         "recorded_by":"U","recorded_on":"2025-06-02T12:00:00","remaining_budget":20}
    ]
    df = pd.DataFrame(exp)
    with pd.ExcelWriter("data/expenses.xlsx", engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Expenses", index=False)

    row, pid, fy, alloc, spent, rem = le.get_budget_info("DeptX","Cat1")
    assert spent == 30
    assert rem == 20

# --- prompt_for_category ---

def test_prompt_for_category(monkeypatch, capsys):
    inputs = iter(["foo","10","2"])
    monkeypatch.setattr("builtins.input", lambda _: next(inputs))
    cat = le.prompt_for_category()
    out = capsys.readouterr().out
    assert "Please enter a number." in out
    assert "Invalid selection." in out
    assert cat == "new_construction"

# --- create_expense_schema ---

def test_create_expense_schema(tmp_env):
    path = le.create_expense_schema()
    assert os.path.exists(path)
    obj = json.load(open(path))
    assert "properties" in obj and "expense_id" in obj["properties"]
    Path(path).unlink()

# --- update_budget_and_log_expense ---

@pytest.fixture(autouse=True)
def patch_excel_tools(monkeypatch):
    import log_expense as module
    monkeypatch.setattr(module, "create_sheets_from_schema", lambda *a,**k: None)
    monkeypatch.setattr(module, "load_workbook", lambda p: openpyxl.load_workbook(p))
    monkeypatch.setattr(module, "save_workbook", lambda wb,p: wb.save(p))

@pytest.fixture
def atomic_setup(tmp_env, budget_only):
    # Pre-create an empty expenses.xlsx so temp_expense_path gets set
    with pd.ExcelWriter("data/expenses.xlsx", engine="openpyxl") as w:
        pd.DataFrame(columns=["expense_id"]).to_excel(w, sheet_name="Expenses", index=False)
    yield

def test_update_and_log_failure_rolls_back(tmp_env, budget_only):
    # monkey-patch append to raise inside update_budget_and_log_expense
    import openpyxl.worksheet.worksheet as wsmod
    orig_append = wsmod.Worksheet.append
    def bad_append(self, *args, **kwargs):
        raise RuntimeError("boom")
    wsmod.Worksheet.append = bad_append

    row, pid, fy, alloc, spent, rem = le.get_budget_info("DeptX","Cat1")
    expense_data = {
        "expense_id": "EXP-FAIL01",
        "department": "DeptX",
        "category": "Cat1",
        "description": "Test fail",
        "date": date.today().isoformat(),
        "recorded_by": "User",
        "amount": 10,
        "project_id": pid,
        "fiscal_year": fy
    }

    with pytest.raises(Exception):
        le.update_budget_and_log_expense(expense_data, row, spent + 10, rem - 10)

    # Ensure no half-written files
    assert not os.path.exists("data/expenses.xlsx") or openpyxl.load_workbook("data/expenses.xlsx").max_row == 1
    wb = openpyxl.load_workbook("data/budget_allocations.xlsx")
    headers = [c.value for c in wb["Allocations"][1]]
    # No spent_amount column if rollback
    assert "spent_amount" not in headers

    # restore
    wsmod.Worksheet.append = orig_append
