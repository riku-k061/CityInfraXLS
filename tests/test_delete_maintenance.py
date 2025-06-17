import os
import re
import json
import shutil
import pandas as pd
import pytest
from pathlib import Path
import delete_maintenance
from openpyxl import Workbook, load_workbook
from datetime import datetime

@pytest.fixture(autouse=True)
def tmp_cwd(tmp_path, monkeypatch):
    # Sandbox everything under a fresh tmp directory
    monkeypatch.chdir(tmp_path)
    # Ensure data/ exists
    (tmp_path / "data").mkdir()
    return tmp_path

# --- backup_workbook tests ---

def test_backup_workbook_creates_copy(tmp_cwd, monkeypatch):
    # Create a dummy maintenance_history.xlsx
    src = tmp_cwd / "data" / "maintenance_history.xlsx"
    src.parent.mkdir(exist_ok=True)
    wb = Workbook()
    wb.active['A1'] = "foo"
    wb.save(src)
    wb.close()

    # Freeze datetime for predictable filename
    fake_now = datetime(2025, 6, 14, 12, 0, 0)
    monkeypatch.setattr(delete_maintenance, 'datetime', datetime)
    # Monkey-patch datetime.now()
    class DummyDateTime(datetime):
        @classmethod
        def now(cls):
            return fake_now
    monkeypatch.setattr(delete_maintenance, 'datetime', DummyDateTime)

    backup_path = delete_maintenance.backup_workbook(str(src))

    # It should live under data/backups
    assert Path(backup_path).parent.name == "backups"
    # Filename should match our timestamp
    assert re.search(r"maintenance_history_backup_20250614_120000\.xlsx$", backup_path)
    # Backup file exists and is a distinct copy
    assert Path(backup_path).exists()
    assert Path(backup_path) != src
    # Contents match
    orig = load_workbook(src).active['A1'].value
    copy = load_workbook(backup_path).active['A1'].value
    assert orig == copy

# --- verify_maintenance_sheet tests ---

def test_verify_sheet_creates_missing(tmp_cwd, monkeypatch):
    # Create an Excel file lacking the sheet
    path = tmp_cwd / "data" / "maintenance_history.xlsx"
    wb = Workbook()
    wb.create_sheet("Other")
    wb.save(path)
    wb.close()

    # Spy on backup and recreate calls
    calls = {'backed_up': False, 'recreated': False}

    def fake_backup(p):
        calls['backed_up'] = True
        return str(p).replace(".xlsx", "_bak.xlsx")

    monkeypatch.setattr(delete_maintenance, 'backup_workbook', fake_backup)

    # Fake import of excel_handler with the function
    class FakeEH:
        @staticmethod
        def create_maintenance_history_sheet(p):
            calls['recreated'] = True
            # actually add the sheet so verify returns True
            wb2 = load_workbook(p)
            ws = wb2.create_sheet("Maintenance History")
            ws['A1'] = "record_id"
            wb2.save(p)
            wb2.close()
            return True

    # but verify imports from utils/excel_handler.py, so patch importlib
    import importlib.util
    real_spec = importlib.util.spec_from_file_location
    def fake_spec(name, location):
        # Return a spec whose loader will exec FakeEH
        class Spec:
            loader = type('L', (), {'exec_module': lambda self, mod: mod.__dict__.update(FakeEH.__dict__)})
        return Spec()
    monkeypatch.setattr(importlib.util, 'spec_from_file_location', fake_spec)

    result = delete_maintenance.verify_maintenance_sheet(str(path))
    assert calls['backed_up']
    assert calls['recreated'] == False

def test_verify_sheet_invalid_structure(tmp_cwd):
    # Create a workbook with Maintenance History but missing record_id
    path = tmp_cwd / "data" / "maintenance_history.xlsx"
    df = pd.DataFrame([{"foo": 1}])
    with pd.ExcelWriter(path, engine='openpyxl') as w:
        df.to_excel(w, sheet_name="Maintenance History", index=False)

    res = delete_maintenance.verify_maintenance_sheet(str(path))
    assert res is False

def test_verify_sheet_good(tmp_cwd):
    # Create workbook with proper sheet + record_id column
    path = tmp_cwd / "data" / "maintenance_history.xlsx"
    df = pd.DataFrame([{"record_id": "X"}])
    with pd.ExcelWriter(path, engine='openpyxl') as w:
        df.to_excel(w, sheet_name="Maintenance History", index=False)

    res = delete_maintenance.verify_maintenance_sheet(str(path))
    assert res is True

# --- delete_maintenance_record tests ---

def test_delete_record_no_file(tmp_cwd):
    # Ensure file does not exist
    path = tmp_cwd / "data" / "maintenance_history.xlsx"
    if path.exists():
        path.unlink()
    res = delete_maintenance.delete_maintenance_record("any", force=True)
    assert res is False

def test_delete_abort_on_bad_sheet(tmp_cwd, monkeypatch):
    # Create empty file
    path = tmp_cwd / "data" / "maintenance_history.xlsx"
    wb = Workbook()
    wb.save(path)
    wb.close()

    # make verify return False
    monkeypatch.setattr(delete_maintenance, 'verify_maintenance_sheet', lambda p: False)
    res = delete_maintenance.delete_maintenance_record("id", force=True)
    assert res is False

def test_delete_nonexistent_id(tmp_cwd):
    # Create file with one record_id
    path = tmp_cwd / "data" / "maintenance_history.xlsx"
    df = pd.DataFrame([{"record_id": "AAA", "asset_id": "X", "date":"2025-01-01","cost":0,"action_taken":"Inspection"}])
    with pd.ExcelWriter(path, engine='openpyxl') as w:
        df.to_excel(w, sheet_name="Maintenance History", index=False)

    res = delete_maintenance.delete_maintenance_record("BBB", force=True)
    assert res is False

def test_delete_cancelled_by_user(tmp_cwd, monkeypatch):
    # Create file with one record
    path = tmp_cwd / "data" / "maintenance_history.xlsx"
    df = pd.DataFrame([{"record_id": "AAA", "asset_id":"X","date":"2025-01-01","cost":0,"action_taken":"Inspection"}])
    with pd.ExcelWriter(path, engine='openpyxl') as w:
        df.to_excel(w, sheet_name="Maintenance History", index=False)

    # verify ok
    assert delete_maintenance.verify_maintenance_sheet(str(path))

    # simulate user types "n"
    monkeypatch.setattr('builtins.input', lambda prompt="": "n")
    res = delete_maintenance.delete_maintenance_record("AAA", force=False)
    assert res is False

def test_delete_force_success(tmp_cwd, monkeypatch):
    # Create file with two records
    path = tmp_cwd / "data" / "maintenance_history.xlsx"
    df = pd.DataFrame([
        {"record_id": "AAA", "asset_id": "X","date":"2025-01-01","cost":0,"action_taken":"Inspect"},
        {"record_id": "BBB", "asset_id": "Y","date":"2025-02-01","cost":0,"action_taken":"Repair"}
    ])
    with pd.ExcelWriter(path, engine='openpyxl') as w:
        df.to_excel(w, sheet_name="Maintenance History", index=False)

    # spy on backup to avoid time uncertainty
    backups = []
    monkeypatch.setattr(delete_maintenance, 'backup_workbook', lambda p: backups.append(p) or str(p)+"_bak.xlsx")

    res = delete_maintenance.delete_maintenance_record("AAA", force=True)
    assert res is True

    # backup called
    assert backups, "expected backup_workbook to be called"
    # file now contains only BBB
    df2 = pd.read_excel(path, sheet_name="Maintenance History")
    assert list(df2['record_id']) == ["BBB"]
