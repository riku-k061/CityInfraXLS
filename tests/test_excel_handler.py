import os
import sys
import json
import pytest
import pandas as pd
from openpyxl import Workbook

# Add parent directory to sys.path to import utils module
sys.path.append(os.path.abspath(os.path.join(os.getcwd(), "utils/..")))

from utils.excel_handler import (
    load_workbook,
    save_workbook,
    init_workbook,
    create_sheets_from_schema,
    create_tasks_sheet
)

@pytest.fixture
def sample_headers():
    return ["ID", "Name", "Location", "Status", "Last Updated"]

@pytest.fixture
def sample_schema():
    return {
        "Road": ["ID", "Name", "Location", "Length", "Width", "Surface Type", "Condition", "Installation Date"],
        "Bridge": ["ID", "Name", "Location", "Length", "Width", "Material", "Condition", "Installation Date"],
        "Park": ["ID", "Name", "Location", "Area", "Facilities", "Condition", "Installation Date"]
    }

@pytest.fixture
def schema_file(tmp_path, sample_schema):
    schema_path = tmp_path / "test_schema.json"
    with open(schema_path, "w") as f:
        json.dump(sample_schema, f)
    return schema_path

@pytest.fixture
def sample_json_schema():
    return {
        "properties": {
            "Contractor ID": {"type": "string"},
            "Name": {"type": "string"},
            "Specialties": {"type": "array"},
            "Regions": {"type": "array"},
            "Rating": {"type": "number"}
        }
    }

@pytest.fixture
def json_schema_file(tmp_path, sample_json_schema):
    schema_path = tmp_path / "contractors_schema.json"
    with open(schema_path, "w") as f:
        json.dump(sample_json_schema, f)
    return schema_path

# --- Legacy tests for classic workbook operations ---

def test_init_workbook_creates_new(tmp_path, sample_headers):
    path = tmp_path / "new.xlsx"
    wb = init_workbook(path, sample_headers)
    assert os.path.exists(path)
    for i, header in enumerate(sample_headers, start=1):
        assert wb.active.cell(row=1, column=i).value == header
    wb.close()

def test_init_workbook_reuses_existing(tmp_path, sample_headers):
    path = tmp_path / "reuse.xlsx"
    wb1 = init_workbook(path, sample_headers)
    wb1.active.cell(row=2, column=1, value="EXISTING")
    wb1.save(path)
    wb1.close()

    wb2 = init_workbook(path, sample_headers)
    assert wb2.active.cell(row=2, column=1).value == "EXISTING"
    wb2.close()

def test_load_workbook_existing(tmp_path):
    path = tmp_path / "load.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="HEADER")
    ws.cell(row=2, column=1, value="DATA")
    wb.save(path)
    wb.close()

    loaded = load_workbook(path)
    ws_loaded = loaded.active
    assert ws_loaded.cell(row=1, column=1).value == "HEADER"
    assert ws_loaded.cell(row=2, column=1).value == "DATA"
    loaded.close()

def test_load_workbook_missing(tmp_path):
    with pytest.raises(Exception):
        load_workbook(tmp_path / "nope.xlsx")

def test_save_workbook_roundtrip(tmp_path):
    path = tmp_path / "roundtrip.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="SAVE")
    save_workbook(wb, path)
    wb.close()

    wb2 = load_workbook(path)
    assert wb2.active.cell(row=1, column=1).value == "SAVE"
    wb2.active.cell(row=1, column=1, value="UPDATED")
    wb2.save(path)
    wb2.close()

    wb3 = load_workbook(path)
    assert wb3.active.cell(row=1, column=1).value == "UPDATED"
    wb3.close()

# --- Updated tests for new schema-based Excel creation ---

def test_create_sheets_from_schema_json(tmp_path, json_schema_file, sample_json_schema):
    path = tmp_path / "contractors.xlsx"
    create_sheets_from_schema(json_schema_file, path)
    df = pd.read_excel(path, sheet_name="contractors")
    assert list(df.columns) == list(sample_json_schema["properties"].keys())

def test_create_sheets_from_schema_custom_sheet(tmp_path, json_schema_file, sample_json_schema):
    path = tmp_path / "custom.xlsx"
    create_sheets_from_schema(json_schema_file, path, sheet_name="MySheet")
    df = pd.read_excel(path, sheet_name="MySheet")
    assert list(df.columns) == list(sample_json_schema["properties"].keys())

def test_create_tasks_sheet(tmp_path):
    path = tmp_path / "tasks.xlsx"
    create_tasks_sheet(output_path=path)
    df = pd.read_excel(path, sheet_name="tasks")
    assert list(df.columns) == ["Task ID", "Incident ID", "Contractor ID", "Assigned At", "Status", "Details"]
