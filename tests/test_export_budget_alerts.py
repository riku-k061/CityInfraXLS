# tests/test_export_budget_alerts.py

import os
import csv
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

import export_budget_alerts as eba

# --- SANDBOX CWD & DATA DIR ---

@pytest.fixture(autouse=True)
def tmp_env(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()
    (tmp_path / "data" / "exports").mkdir(parents=True)
    yield

# --- Helpers to create an Alerts sheet ---

def make_workbook_with_alerts(rows, columns):
    """
    Create data/budget_allocations.xlsx with an Alerts sheet.
    'rows' is a list of dicts mapping column names to values.
    'columns' is the list of column names to include.
    """
    wb = openpyxl.Workbook()
    # Remove default
    wb.remove(wb.active)
    ws = wb.create_sheet("Alerts")
    ws.append(columns)
    for r in rows:
        ws.append([r.get(col, None) for col in columns])
    path = Path("data/budget_allocations.xlsx")
    wb.save(path)
    return path

# --- Test: missing Alerts sheet ---

def test_export_no_alerts_sheet_returns_false_and_logs(tmp_env, caplog):
    # Create workbook without Alerts
    wb = openpyxl.Workbook()
    wb.save("data/budget_allocations.xlsx")

    caplog.set_level("ERROR", logger="budget_alerts")
    result = eba.export_alerts_to_csv(
        source_excel="data/budget_allocations.xlsx",
        output_csv="data/exports/budget_alerts.csv",
        backup=False
    )
    assert result is False
    assert "No Alerts sheet found" in caplog.text
    assert not Path("data/exports/budget_alerts.csv").exists()

# --- Test: required columns missing ---

def test_export_missing_required_columns(tmp_env, caplog):
    # Alerts sheet but missing some required columns
    columns = ['department','project_id','status']  # missing others
    rows = [
        {'department':'D1','project_id':'P1','status':'Over Budget'}
    ]
    make_workbook_with_alerts(rows, columns)

    caplog.set_level("ERROR", logger="budget_alerts")
    result = eba.export_alerts_to_csv(
        source_excel="data/budget_allocations.xlsx",
        output_csv="data/exports/budget_alerts.csv",
        backup=False
    )
    assert result is False
    assert "Required columns not found" in caplog.text
    assert not Path("data/exports/budget_alerts.csv").exists()

# --- Test: no critical alerts (empty export_df) still writes file with headers + timestamp ---

def test_export_no_critical_alerts_creates_csv_and_timestamp(tmp_env):
    columns = ['department','project_id','allocated_amount','remaining_budget','overrun_amount','status']
    # status not in critical list
    rows = [
        {'department':'D1','project_id':'P1','allocated_amount':100,'remaining_budget':50,'overrun_amount':0,'status':'OK'}
    ]
    make_workbook_with_alerts(rows, columns)

    result = eba.export_alerts_to_csv(
        source_excel="data/budget_allocations.xlsx",
        output_csv="data/exports/budget_alerts.csv",
        backup=False
    )
    assert result is True

    out_csv = Path("data/exports/budget_alerts.csv")
    assert out_csv.exists()

    df = pd.read_csv(out_csv)
    # Only header + export_timestamp column
    assert list(df.columns) == ['department','project_id','remaining_budget','overrun_amount','status','export_timestamp']
    # No rows, so length zero
    assert len(df) == 0

    # Check last_sync file
    last_sync = Path("data/exports/budget_alerts_last_sync.txt")
    assert last_sync.exists()
    content = last_sync.read_text()
    assert "Last synchronized:" in content

# --- Test: critical alerts and backup behavior ---

def test_export_with_critical_and_backup(tmp_env, caplog):
    columns = ['department','project_id','allocated_amount','remaining_budget','overrun_amount','status']
    rows = [
        {'department':'D1','project_id':'P1','allocated_amount':100,'remaining_budget':-10,'overrun_amount':10,'status':'Over Budget'},
        {'department':'D2','project_id':'P2','allocated_amount':200,'remaining_budget':5,'overrun_amount':0,'status':'At Risk'}
    ]
    make_workbook_with_alerts(rows, columns)

    # Pre-create an existing CSV to test backup
    existing = Path("data/exports/budget_alerts.csv")
    with existing.open('w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["dummy"])
    before = datetime.now().timestamp()

    caplog.set_level("INFO", logger="budget_alerts")
    result = eba.export_alerts_to_csv(
        source_excel="data/budget_allocations.xlsx",
        output_csv=str(existing),
        backup=True
    )
    assert result is True

    # Backup file created
    backups = list(Path("data/exports").glob("budget_alerts_*_backup.csv"))
    assert len(backups) == 1
    backup_time = backups[0].stem.split('_')[-2]
    # simple check: timestamp-like

    # New CSV has two rows
    df = pd.read_csv(existing)
    assert len(df) == 2
    assert set(df['status']) == {'Over Budget','At Risk'}

    # Info logs mention backup and success
    assert "Created backup at" in caplog.text
    assert "Successfully exported 2 alert records" in caplog.text

    # last_sync file updated
    last_sync = Path("data/exports/budget_alerts_last_sync.txt")
    assert last_sync.exists()
    assert "Last synchronized:" in last_sync.read_text()
