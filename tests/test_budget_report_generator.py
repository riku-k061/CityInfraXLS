# tests/test_budget_report_generator.py

import os
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

import budget_report_generator as brg

# --- FIXTURE: SANDBOX CWD & DATA DIR ---

@pytest.fixture(autouse=True)
def tmp_env(tmp_path, monkeypatch):
    """
    Run each test in its own temp directory, with data/ and reports/ subfolders.
    """
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()
    (tmp_path / "reports").mkdir()
    yield

# --- format_currency tests ---

def test_format_currency_na():
    assert brg.format_currency(float("nan")) == "$0.00"

def test_format_currency_values():
    assert brg.format_currency(0) == "$0.00"
    assert brg.format_currency(123.456) == "$123.46"
    assert brg.format_currency(5) == "$5.00"

# --- generate_budget_report tests ---

def write_budget_and_expenses(budgets, expenses):
    """
    Write data/budget_allocations.xlsx and data/expenses.xlsx from lists of dicts,
    ensuring the correct columns exist even if lists are empty.
    """
    # Budget sheet must have these columns:
    budget_cols = ["department","project_id","category",
                   "allocation_date","allocated_amount","fiscal_year"]
    df_bud = pd.DataFrame(budgets, columns=budget_cols)
    with pd.ExcelWriter("data/budget_allocations.xlsx", engine="openpyxl") as w:
        df_bud.to_excel(w, sheet_name="Allocations", index=False)
    # Expense sheet must have these columns:
    expense_cols = ["expense_id","department","project_id","date","category",
                    "amount","description","recorded_by","fiscal_year"]
    df_exp = pd.DataFrame(expenses, columns=expense_cols)
    with pd.ExcelWriter("data/expenses.xlsx", engine="openpyxl") as w:
        df_exp.to_excel(w, sheet_name="Expenses", index=False)

def test_no_allocations_raises():
    # Empty budget, but with proper header row
    write_budget_and_expenses([], [])
    with pytest.raises(ValueError) as exc:
        brg.generate_budget_report(output_path="reports/out.xlsx", fiscal_year="2025-2026")
    assert "No budget allocations found" in str(exc.value)

def test_generate_report_basic():
    # Sample data for FY 2025-2026
    budgets = [
        {"department":"D1","project_id":"P1","category":"CatA",
         "allocation_date":"2025-01-01","allocated_amount":100,"fiscal_year":"2025-2026"},
        {"department":"D2","project_id":"P2","category":"CatB",
         "allocation_date":"2025-02-01","allocated_amount":200,"fiscal_year":"2025-2026"}
    ]
    expenses = [
        {"expense_id":"E1","department":"D1","project_id":"P1","date":"2025-03-01","category":"CatA",
         "amount": 30,"description":"desc","recorded_by":"U","fiscal_year":"2025-2026"}
    ]
    write_budget_and_expenses(budgets, expenses)

    out_path = "reports/test_report.xlsx"
    ret = brg.generate_budget_report(output_path=out_path, fiscal_year="2025-2026")
    assert ret == out_path
    assert Path(out_path).exists()

    wb = openpyxl.load_workbook(out_path)
    # Sheets
    assert set(wb.sheetnames) >= {"Summary","Department Details","Alerts"}

    # Summary sheet title
    ws = wb["Summary"]
    assert ws["A1"].value.startswith("Budget Summary Report - Fiscal Year 2025-2026")

    # Summary header row at row 2
    assert [c.value for c in ws[2][:6]] == ["Department","Allocated","Spent","Remaining","% Used","Status"]

    # Check D1 row values
    for row in ws.iter_rows(min_row=3, max_col=6, values_only=True):
        if row[0] == "D1":
            assert row[1] == "$100.00"
            assert row[2] == "$30.00"
            assert row[3] == "$70.00"
            assert row[4] == "30.0%"
            assert row[5] == "ACTIVE"
            break
    else:
        pytest.fail("D1 not found in summary")

    # Alerts sheet should contain "No budget alerts"
    ws_alerts = wb["Alerts"]
    texts = [c.value for c in ws_alerts["A"]]
    assert any("No budget alerts" in str(x) for x in texts)

def test_generate_report_auto_fiscal_year():
    # Write only budgets (no expenses) for FY 2025-2026
    budgets = [
        {"department":"D1","project_id":"P1","category":"CatA",
         "allocation_date":"2025-01-01","allocated_amount":50,"fiscal_year":"2025-2026"}
    ]
    write_budget_and_expenses(budgets, [])  # empty expenses but correct headers

    out = "reports/auto.xlsx"
    ret = brg.generate_budget_report(output_path=out, fiscal_year=None)
    assert ret == out
    # File created
    assert Path(out).exists()
    # Title reflects chosen fiscal year
    wb = openpyxl.load_workbook(out)
    title = wb["Summary"]["A1"].value
    assert "Fiscal Year 2025-2026" in title
