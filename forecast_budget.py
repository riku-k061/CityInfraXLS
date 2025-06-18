# forecast_budget.py
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.sparkline import SparklineGroup, Sparkline
import json
from datetime import datetime
import os
from pathlib import Path

def load_schema():
    """Load budget allocation schema for validation"""
    with open('budget_allocation_schema.json', 'r') as f:
        return json.load(f)

def calculate_year_end_forecast(data_path='data/budget_allocations.xlsx'):
    """
    Calculate projected year-end spending based on historical monthly averages
    and write the forecast back to the allocations file with sparkline visualizations.
    
    Parameters:
        data_path: Path to the Excel file containing budget data
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Load allocations data
        allocations_df = pd.read_excel(data_path, sheet_name='Allocations')
        
        # Load actuals data (monthly spending)
        actuals_df = pd.read_excel(data_path, sheet_name='Actuals')
        
        # Current month (1-12)
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        # Create result dataframes
        forecast_results = []
        
        # Group actuals by department and project_id
        grouped = actuals_df.groupby(['department', 'project_id'])
        
        for (dept, proj_id), group in grouped:
            # Get allocated budget for this department/project
            allocated = allocations_df[
                (allocations_df['department'] == dept) & 
                (allocations_df['project_id'] == proj_id)
            ]['allocated_amount'].iloc[0] if not allocations_df[
                (allocations_df['department'] == dept) & 
                (allocations_df['project_id'] == proj_id)
            ].empty else 0
            
            # Get monthly spending data
            monthly_spend = group.pivot_table(
                index=['department', 'project_id'], 
                columns='month', 
                values='amount', 
                aggfunc='sum'
            ).fillna(0).iloc[0]
            
            # Calculate spending to date
            spend_to_date = sum(monthly_spend.get(i, 0) for i in range(1, current_month + 1))
            
            # Calculate monthly average spending
            months_with_data = sum(1 for i in range(1, current_month + 1) if monthly_spend.get(i, 0) > 0)
            avg_monthly_spend = spend_to_date / max(months_with_data, 1)
            
            # Project remaining months
            remaining_months = 12 - current_month
            projected_additional_spend = avg_monthly_spend * remaining_months
            
            # Calculate projected year-end total
            projected_total = spend_to_date + projected_additional_spend
            
            # Calculate percent of allocation
            percent_used = (spend_to_date / allocated * 100) if allocated > 0 else 0
            projected_percent = (projected_total / allocated * 100) if allocated > 0 else 0
            
            # Determine status based on projected spend vs allocation
            if projected_percent > 110:
                status = "Over Budget"
            elif projected_percent > 98:
                status = "At Risk"
            elif projected_percent < 75:
                status = "Underspend"
            else:
                status = "On Track"
                
            # Get monthly spending for sparkline
            monthly_values = [monthly_spend.get(i, 0) for i in range(1, 13)]
                
            # Store result
            forecast_results.append({
                'department': dept,
                'project_id': proj_id,
                'allocated_amount': allocated,
                'spend_to_date': spend_to_date,
                'percent_used': percent_used,
                'average_monthly_spend': avg_monthly_spend,
                'projected_year_end': projected_total,
                'projected_percent': projected_percent,
                'status': status,
                'monthly_values': monthly_values  # For sparkline chart
            })
        
        # Create DataFrame from results
        forecast_df = pd.DataFrame(forecast_results)
        
        # Save to Excel with formatting and sparklines
        create_forecast_sheet(data_path, forecast_df)
        
        print(f"Budget forecasts have been calculated and saved to {data_path}.")
        return True
        
    except Exception as e:
        print(f"Error calculating budget forecasts: {e}")
        return False

def create_forecast_sheet(excel_path, forecast_df):
    """
    Creates or updates the Forecasts sheet in the Excel file and adds sparklines
    to visualize the spending trajectory.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_path)
    
    # Create or get the Forecasts sheet
    if "Forecasts" in workbook.sheetnames:
        # Remove existing sheet to recreate it
        workbook.remove(workbook["Forecasts"])
    
    forecast_sheet = workbook.create_sheet("Forecasts")
    
    # Add title
    forecast_sheet['A1'] = "BUDGET FORECASTS"
    forecast_sheet['A1'].font = Font(size=16, bold=True)
    forecast_sheet.merge_cells('A1:J1')
    forecast_sheet['A1'].alignment = Alignment(horizontal='center')
    
    # Add current date
    forecast_sheet['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    forecast_sheet['A2'].font = Font(italic=True)
    forecast_sheet.merge_cells('A2:J2')
    
    # Add headers
    headers = [
        "Department", "Project ID", "Allocated Budget", "Spend to Date", 
        "% Used", "Monthly Avg", "Year-End Projection", "Proj %", "Status", "Trend"
    ]
    for col, header in enumerate(headers, start=1):
        cell = forecast_sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add monthly data in hidden columns for sparklines
    for col, month in enumerate(range(1, 13), start=11):
        cell = forecast_sheet.cell(row=4, column=col)
        cell.value = f"Month {month}"
        # Hide these columns as they're just for sparkline data
        forecast_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True
    
    # Add data rows
    for i, row_data in enumerate(forecast_df.itertuples(), start=5):
        # Regular data columns
        forecast_sheet.cell(row=i, column=1).value = row_data.department
        forecast_sheet.cell(row=i, column=2).value = row_data.project_id
        forecast_sheet.cell(row=i, column=3).value = row_data.allocated_amount
        forecast_sheet.cell(row=i, column=3).number_format = '#,##0.00'
        
        forecast_sheet.cell(row=i, column=4).value = row_data.spend_to_date
        forecast_sheet.cell(row=i, column=4).number_format = '#,##0.00'
        
        forecast_sheet.cell(row=i, column=5).value = row_data.percent_used / 100  # As decimal for percentage format
        forecast_sheet.cell(row=i, column=5).number_format = '0.00%'
        
        forecast_sheet.cell(row=i, column=6).value = row_data.average_monthly_spend
        forecast_sheet.cell(row=i, column=6).number_format = '#,##0.00'
        
        forecast_sheet.cell(row=i, column=7).value = row_data.projected_year_end
        forecast_sheet.cell(row=i, column=7).number_format = '#,##0.00'
        
        forecast_sheet.cell(row=i, column=8).value = row_data.projected_percent / 100  # As decimal for percentage format
        forecast_sheet.cell(row=i, column=8).number_format = '0.00%'
        
        status_cell = forecast_sheet.cell(row=i, column=9)
        status_cell.value = row_data.status
        
        # Color code status
        if row_data.status == "Over Budget":
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        elif row_data.status == "At Risk":
            status_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
        elif row_data.status == "Underspend":
            status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        else:
            status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        
        # Add monthly values in hidden columns for sparkline data
        for j, monthly_value in enumerate(row_data.monthly_values, start=0):
            forecast_sheet.cell(row=i, column=11 + j).value = monthly_value
    
    # Add sparklines to the data
    # We need to define the sparklines after all data is entered
    sparklines = []
    for i in range(5, 5 + len(forecast_df)):
        # Define the source data range (the hidden columns with monthly data)
        # Starts at column 11 (Month 1) and goes through column 22 (Month 12)
        data_range = f'{openpyxl.utils.get_column_letter(11)}{i}:{openpyxl.utils.get_column_letter(22)}{i}'
        
        # Define where the sparkline should be placed (in column J)
        location = f'J{i}'
        
        # Create a Sparkline object for this row
        sparkline = Sparkline(data_range, location)
        sparklines.append(sparkline)
    
    # Create a SparklineGroup containing all the sparklines
    # Colors are: first marker, last marker, high marker, low marker, negative
    sparkline_group = SparklineGroup(
        sparklines=sparklines, 
        type='line', 
        displayEmptyCellsAs='gap', 
        markers=True, 
        high=True, 
        low=True, 
        first=True, 
        last=True,
        colorSeries='1F77B4',  # Blue for the line
        colorFirst='0000FF',   # Blue for first marker
        colorLast='FF0000',    # Red for last marker
        colorHigh='00FF00',    # Green for high marker
        colorLow='FF0000'      # Red for low marker
    )
    
    # Add the sparkline group to the worksheet
    forecast_sheet.add_sparkline_group(sparkline_group)
    
    # Set column widths
    for col in range(1, 11):
        column_letter = openpyxl.utils.get_column_letter(col)
        forecast_sheet.column_dimensions[column_letter].width = 15
    
    # Make the Trend column wider for the sparkline
    forecast_sheet.column_dimensions['J'].width = 20
    
    # Save the workbook
    workbook.save(excel_path)

if __name__ == "__main__":
    calculate_year_end_forecast()