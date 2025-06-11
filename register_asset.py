# register_asset.py

import os
import json
import uuid
import datetime
from pathlib import Path

# Import our custom Excel utilities
from utils.excel_handler import load_workbook, save_workbook, create_sheets_from_schema

# Constants
SCHEMA_PATH = 'asset_schema.json'
ASSETS_PATH = 'data/assets.xlsx'
LOG_PATH = 'data/asset_log.xlsx'
LOG_HEADERS = ['Timestamp', 'Asset ID', 'Asset Type', 'Action', 'Details']

def ensure_data_directory():
    """Ensure the data directory exists"""
    Path('data').mkdir(exist_ok=True)
    
def load_schema():
    """Load the asset schema from JSON"""
    with open(SCHEMA_PATH, 'r') as f:
        return json.load(f)

def validate_input(prompt, allow_empty=False):
    """Prompt for input and validate it's not empty (unless allowed)"""
    while True:
        value = input(prompt).strip()
        if value or allow_empty:
            return value
        print("This field cannot be empty. Please try again.")

def register_asset():
    """Main function to register a new asset"""
    # Ensure data directory exists
    ensure_data_directory()
    
    # Load asset schema
    schema = load_schema()
    
    # Prepare workbooks
    create_sheets_from_schema(SCHEMA_PATH, ASSETS_PATH)
    
    # Initialize asset log file if it doesn't exist
    if not os.path.exists(LOG_PATH):
        log_wb = load_workbook(LOG_PATH) if os.path.exists(LOG_PATH) else None
        if log_wb is None:
            from openpyxl import Workbook
            log_wb = Workbook()
            log_ws = log_wb.active
            log_ws.title = "Asset Log"
            for col_idx, header in enumerate(LOG_HEADERS, start=1):
                log_ws.cell(row=1, column=col_idx, value=header)
            log_wb.save(LOG_PATH)
    
    # Display available asset types
    asset_types = list(schema.keys())
    print("\nAvailable asset types:")
    for i, asset_type in enumerate(asset_types, 1):
        print(f"{i}. {asset_type}")
    
    # Get asset type choice
    while True:
        choice = validate_input("\nSelect asset type (enter number): ")
        try:
            choice_idx = int(choice) - 1
            if 0 <= choice_idx < len(asset_types):
                asset_type = asset_types[choice_idx]
                break
            else:
                print(f"Please enter a number between 1 and {len(asset_types)}")
        except ValueError:
            print("Please enter a valid number")
    
    # Generate UUID for the asset
    asset_id = str(uuid.uuid4())
    print(f"\nGenerated Asset ID: {asset_id}")
    
    # Collect data for each field in the schema
    asset_data = {'ID': asset_id}
    fields = schema[asset_type]
    
    print(f"\nEnter details for {asset_type}:")
    for field in fields:
        if field != 'ID':  # Skip ID since we already generated it
            value = validate_input(f"{field}: ")
            asset_data[field] = value
    
    # Load assets workbook and append data
    assets_wb = load_workbook(ASSETS_PATH)
    assets_ws = assets_wb[asset_type]
    
    # Find the next empty row
    next_row = assets_ws.max_row + 1
    for col_idx, field in enumerate(fields, start=1):
        assets_ws.cell(row=next_row, column=col_idx, value=asset_data.get(field, ''))
    
    # Save the updated assets workbook
    save_workbook(assets_wb, ASSETS_PATH)
    
    # Log the registration
    log_wb = load_workbook(LOG_PATH)
    log_ws = log_wb.active
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = [timestamp, asset_id, asset_type, 'REGISTER', 'New asset registered']
    
    next_row = log_ws.max_row + 1
    for col_idx, value in enumerate(log_entry, start=1):
        log_ws.cell(row=next_row, column=col_idx, value=value)
    
    # Save the updated log workbook
    save_workbook(log_wb, LOG_PATH)
    
    print(f"\nSuccess! {asset_type} with ID {asset_id} has been registered.")
    print(f"A log entry has been added to {LOG_PATH}")

if __name__ == "__main__":
    print("=== CityInfraXLS - Asset Registration ===")
    try:
        register_asset()
    except KeyboardInterrupt:
        print("\nOperation cancelled.")
    except Exception as e:
        print(f"\nAn error occurred: {str(e)}")