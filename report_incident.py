# report_incident.py
"""
Interactive script for reporting new infrastructure incidents.
Captures incident details, calculates SLA deadline, and logs incidents.
"""

import os
import uuid
import logging
import datetime
import openpyxl
import sys
from pathlib import Path
from typing import Dict, Any

# Add project root to path to enable imports from other modules
sys.path.insert(0, str(Path(__file__).resolve().parent))

# Import project modules
from validate_severity_matrix import validate_severity_matrix
from utils.incident_handler import create_incident_sheet

# Configure logging
logging.basicConfig(
    filename='cityinfraxls.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('CityInfraXLS')

def load_severity_matrix() -> Dict[str, Dict[str, Any]]:
    """
    Loads and validates the severity matrix.
    
    Returns:
        Dict containing severity levels and their details
    """
    try:
        return validate_severity_matrix('severity_matrix.json')
    except (FileNotFoundError, ValueError) as e:
        logger.error(f"Failed to load severity matrix: {e}")
        print(f"ERROR: Could not load severity matrix - {e}")
        sys.exit(1)

def ensure_incident_sheet() -> str:
    """
    Ensures that the incidents sheet exists, creating it if necessary.
    
    Returns:
        Path to the incidents Excel file
    """
    incident_file = 'data/incidents.xlsx'
    if not os.path.exists(incident_file):
        create_incident_sheet(incident_file)
        logger.info(f"Created new incident tracking sheet at {incident_file}")
    return incident_file

def get_user_input(prompt: str, required: bool = True) -> str:
    """
    Get user input with optional validation for required fields.
    
    Args:
        prompt: Text to display to the user
        required: Whether the field is required
    
    Returns:
        User's input
    """
    while True:
        value = input(prompt).strip()
        if value or not required:
            return value
        print("This field is required. Please enter a value.")

def select_severity(severity_matrix: Dict[str, Dict[str, Any]]) -> str:
    """
    Allows the user to select a severity level.
    
    Args:
        severity_matrix: Dictionary of available severity levels
    
    Returns:
        Selected severity level
    """
    print("\nAvailable severity levels:")
    for idx, (level, details) in enumerate(severity_matrix.items(), 1):
        print(f"{idx}. {level} (SLA: {details['hours']} hours) - {details['description']}")
    
    while True:
        try:
            choice = int(get_user_input(f"\nSelect severity level (1-{len(severity_matrix)}): "))
            if 1 <= choice <= len(severity_matrix):
                # Convert choice to the actual severity level name
                return list(severity_matrix.keys())[choice - 1]
            print(f"Please enter a number between 1 and {len(severity_matrix)}")
        except ValueError:
            print("Please enter a valid number")

def calculate_sla_deadline(severity_level: str, severity_matrix: Dict[str, Dict[str, Any]]) -> datetime.datetime:
    """
    Calculates the SLA deadline based on severity level.
    
    Args:
        severity_level: The selected severity level
        severity_matrix: Dictionary containing severity details
    
    Returns:
        Datetime object representing the SLA deadline
    """
    current_time = datetime.datetime.now()
    sla_hours = severity_matrix[severity_level]["hours"]
    deadline = current_time + datetime.timedelta(hours=sla_hours)
    return deadline

def append_incident(incident_data: Dict[str, Any], file_path: str) -> None:
    """
    Appends a new incident to the incidents Excel file.
    
    Args:
        incident_data: Dictionary containing incident details
        file_path: Path to the incidents Excel file
    """
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Determine the next row
    next_row = ws.max_row + 1
    
    # Add the incident data
    ws.cell(row=next_row, column=1).value = incident_data["incident_id"]
    ws.cell(row=next_row, column=2).value = incident_data["asset_id"]
    ws.cell(row=next_row, column=3).value = incident_data["reporter"]
    ws.cell(row=next_row, column=4).value = incident_data["type"]
    ws.cell(row=next_row, column=5).value = incident_data["severity"]
    ws.cell(row=next_row, column=6).value = incident_data["reported_at"]
    ws.cell(row=next_row, column=7).value = incident_data["sla_deadline"]
    ws.cell(row=next_row, column=8).value = incident_data["status"]
    
    # Save the workbook
    wb.save(file_path)

def main():
    """Main function to run the incident reporting script."""
    print("=== CityInfraXLS: Report Infrastructure Incident ===")
    
    # Load severity matrix
    severity_matrix = load_severity_matrix()
    
    # Ensure incidents sheet exists
    incident_file = ensure_incident_sheet()
    
    # Collect incident details
    asset_id = get_user_input("Asset ID: ")
    issue_type = get_user_input("Issue Type: ")
    reporter = get_user_input("Reporter Name: ")
    severity = select_severity(severity_matrix)
    
    # Generate incident ID
    incident_id = str(uuid.uuid4())
    
    # Get current time and calculate deadline
    current_time = datetime.datetime.now()
    deadline = calculate_sla_deadline(severity, severity_matrix)
    
    # Create incident data
    incident_data = {
        "incident_id": incident_id,
        "asset_id": asset_id,
        "reporter": reporter,
        "type": issue_type,
        "severity": severity,
        "reported_at": current_time,
        "sla_deadline": deadline,
        "status": "Open"  # Default status for new incidents
    }
    
    # Append to spreadsheet
    try:
        append_incident(incident_data, incident_file)
        print(f"\nIncident reported successfully!")
        print(f"Incident ID: {incident_id}")
        print(f"Severity: {severity}")
        print(f"SLA Deadline: {deadline.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Log the incident
        logger.info(f"New incident reported - ID: {incident_id}, Severity: {severity}, Deadline: {deadline.strftime('%Y-%m-%d %H:%M:%S')}")
        
    except Exception as e:
        logger.error(f"Failed to report incident: {e}")
        print(f"ERROR: Could not report incident - {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()